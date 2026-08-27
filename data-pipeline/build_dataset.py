# -*- coding: utf-8 -*-
"""
원본 소스 -> Data_Preprocessing_Template.ts (City/Artist/Place) 3파일 스키마로 정제/병합.

입력 소스:
    1. Kpop_Tour_Spots_Combined.xlsx ("All K-pop Tour Spots" 시트) - 가장 깨끗한 구조화 데이터
    2. stara_thisismykorea.json - 구조화됨 (visitkorea.or.kr)
    3. stara_places.json - 기사 목차형 텍스트라 노이즈 필터링 필요
    4. raw_data_정인지/*.json - 팀원이 블로그 등에서 직접 확인하고 정리한 소스
    5. Kpop_Idol_Hometowns_filtered.xlsx + 아이돌_출신지_*.csv - 아이돌 출신지(고향).
       현재 Place 스키마는 위경도가 있는 "장소"만 다루도록 되어 있어서, 도시 단위인
       출신지 데이터는 참고용으로만 집계하고 최종 출력엔 포함하지 않음.

출력 (기본: 전체 도시 통합. --city=incheon 처럼 도시 하나만 뽑을 수도 있음).
파일명 규칙(2026-08-25 확정): "generated_{종류}_{도시}_{작성자}.{확장자}" -
종류가 맨 앞, 도시/작성자는 뒤, 구분자는 언더바(생략된 값은 그냥 빠짐):
    preprocessed/generated_cities_{city}_{owner}.json  - {city_id, city_name_ko(축약형), city_name_en}
    preprocessed/generated_artists_{city}_{owner}.json - {artist_id, artist_name_ko, artist_name_en}
    preprocessed/generated_places_{city}_{owner}.json  - Place[] (Data_Preprocessing_Template.ts 참고)
    preprocessed/generated_places_needs_review_{city}_{owner}.csv - 이름/출처/
      relation_text 중 하나라도 비어서 사람이 봐야 하는 행
이 파일들은 전부 검수 전 후보다 - 검수 + fill_hours.py로 영업시간까지 끝나
병합된 도시만 preprocessed/final/{cities,artists,places}.json(정본)에 들어간다
(사람이 직접 병합 - 아직 자동화 안 됨).

실행:
    python build_dataset.py                        # 지오코딩 없이 (빠름, 반복 작업용)
    python build_dataset.py --geocode               # 좌표/한글 상호명/카테고리까지 채움
    python build_dataset.py --geocode --city=incheon # 한 도시만 출력
    python build_dataset.py --city=incheon --owner=정인지 # 도시 분담 시 담당자 이름을 파일명에 표시
      -> preprocessed/generated_cities_incheon_정인지.json 등

주의:
    - place_id/artist_id는 최종 조립 단계(7단계)에서만 부여됨. place_category를
      ID의 "맥락" 자리에 쓰고, relation_text에서 특정 멤버가 감지되면
      "{그룹}-{멤버}" 형태로 자동 분리를 시도함(완벽하지 않음 - 코드 주석 참고)
    - source_url이 실제 링크(http/https)가 아니어도, relation_text 자체가
      "인물+구체적 콘텐츠명"을 담은 추적 가능한 인용이면 인정한다(2026-08-25
      정책, has_credible_citation() 참고) - 그 외에는 전부 "PENDING"으로 통일
    - artist_name_ko는 확실히 아는 것만 채움(GROUP_KO_NAMES/MEMBER_KO_NAMES) -
      모르면 영문명을 임시로 넣고 needs_ko_name으로 표시 (틀린 번역보다 안전)
    - 좌표가 없는 장소는 최종 places.json에서 제외됨(필수 필드라 null 불가)
    - status는 전부 "draft"로 시작 - verified/published 승격은 사람이 실제
      출처를 확인한 뒤 수동으로 올려야 함(자동화 대상 아님)
"""

import csv
import json
import os
import re
import sys
import time
import difflib
from collections import OrderedDict

import openpyxl
import requests

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "from drive", "raw_data_곽채원")
if not os.path.isdir(RAW_DIR):
    RAW_DIR = os.path.join(BASE_DIR, "raw_data_곽채원")  # 예전 위치 폴백
OUT_DIR = os.path.join(BASE_DIR, "preprocessed")

HANGUL_RE = re.compile(r"[가-힣]")

# ============================================================
# 1. 지역(광역시/도) 매핑 - 17개 단위로 통일
# ============================================================

# 정규화 키(소문자, 공백/하이픈 제거) -> (city_id, city_name_ko, city_name_en)
PROVINCE_TABLE = [
    ("seoul", "서울특별시", "Seoul"),
    ("busan", "부산광역시", "Busan"),
    ("daegu", "대구광역시", "Daegu"),
    ("incheon", "인천광역시", "Incheon"),
    ("gwangju", "광주광역시", "Gwangju"),
    ("daejeon", "대전광역시", "Daejeon"),
    ("ulsan", "울산광역시", "Ulsan"),
    ("sejong", "세종특별자치시", "Sejong"),
    ("gyeonggi", "경기도", "Gyeonggi"),
    ("gangwon", "강원특별자치도", "Gangwon"),
    ("chungcheongbuk", "충청북도", "Chungcheongbuk"),
    ("chungbuk", "충청북도", "Chungcheongbuk"),
    ("chungcheongnam", "충청남도", "Chungcheongnam"),
    ("chungnam", "충청남도", "Chungcheongnam"),
    ("jeollabuk", "전라북도", "Jeollabuk"),
    ("jeonbuk", "전라북도", "Jeollabuk"),
    ("jeollanam", "전라남도", "Jeollanam"),
    ("jeonnam", "전라남도", "Jeollanam"),
    ("gyeongsangbuk", "경상북도", "Gyeongsangbuk"),
    ("gyeongbuk", "경상북도", "Gyeongsangbuk"),
    ("gyeongsangnam", "경상남도", "Gyeongsangnam"),
    ("gyeongnam", "경상남도", "Gyeongsangnam"),
    ("jeju", "제주특별자치도", "Jeju"),
]

# 구/시/군 이름 -> 소속 광역시/도 정규화 키 (영문 주소에서 상위 지역명이
# 아예 안 나오고 구/시 이름만 있는 경우 보강용. 데이터에 실제 등장하는 이름만 등록)
DISTRICT_TO_PROVINCE = {
    # Seoul 25개 구
    "jung": "seoul", "jongno": "seoul", "yongsan": "seoul", "seongdong": "seoul",
    "gwangjin": "seoul", "dongdaemun": "seoul", "jungnang": "seoul", "seongbuk": "seoul",
    "gangbuk": "seoul", "dobong": "seoul", "nowon": "seoul", "eunpyeong": "seoul",
    "seodaemun": "seoul", "mapo": "seoul", "yangcheon": "seoul", "gangseo": "seoul",
    "guro": "seoul", "geumcheon": "seoul", "yeongdeungpo": "seoul", "dongjak": "seoul",
    "gwanak": "seoul", "seocho": "seoul", "gangnam": "seoul", "songpa": "seoul",
    "gangdong": "seoul",
    "myeongdong": "seoul", "hongdae": "seoul", "apgujeong": "seoul", "cheongdam": "seoul",
    "itaewon": "seoul", "insadong": "seoul", "hannam": "seoul", "seongsu": "seoul",
    "hanam": "seoul",  # xlsx 소스에 "Hanam, Seoul"로 등장 (구 단위 취급됨, 원자료 그대로)
    # Busan
    "haeundae": "busan", "gijang": "busan", "dongnae": "busan", "suyeong": "busan",
    "sasang": "busan", "saha": "busan",
    # Gyeonggi 주요 시
    "goyang": "gyeonggi", "gwacheon": "gyeonggi", "bucheon": "gyeonggi", "guri": "gyeonggi",
    "siheung": "gyeonggi", "ansan": "gyeonggi", "namyangju": "gyeonggi", "yongin": "gyeonggi",
    "anyang": "gyeonggi", "hwaseong": "gyeonggi", "gunpo": "gyeonggi", "seongnam": "gyeonggi",
    "suwon": "gyeonggi", "gimpo": "gyeonggi", "uijeongbu": "gyeonggi",
    # 기타 (경상/전라/강원/제주 등 시 단위)
    "suncheon": "jeollanam", "changwon": "gyeongsangnam", "sinan": "jeollanam",
    "chungju": "chungcheongbuk", "iksan": "jeollabuk", "jeonju": "jeollabuk",
    "yangsan": "gyeongsangnam", "pohang": "gyeongsangbuk", "namhae": "gyeongsangnam",
    "sacheon": "gyeongsangnam", "chuncheon": "gangwon", "gangneung": "gangwon",
    "gyeongju": "gyeongsangbuk", "jejucity": "jeju", "dalseo": "daegu",
    "bupyeong": "incheon", "namdong": "incheon", "geumjeong": "busan", "buk": None,  # 여러 도시에 중복 -> 단독 판단 불가
}

NON_KOREA_KEYWORDS = {
    "china", "japan", "uae", "dubai", "usa", "thailand", "hokkaido", "dongyang",
    "vietnam", "singapore", "taiwan", "hong kong", "malaysia", "philippines",
}


def _norm_key(s):
    # 영문은 소문자 알파벳만, 한글은 완성형 음절만 남김.
    # (주의: 원래 [^a-z]만 걸렀더니 한글 문자열은 전부 지워져서 빈 문자열이 되고,
    #  결과적으로 한글 주소는 단 하나도 지역 매칭이 안 되는 버그가 있었음)
    return re.sub(r"[^a-z가-힣]", "", s.lower())


PROVINCE_LOOKUP = {}
for key, ko, en in PROVINCE_TABLE:
    PROVINCE_LOOKUP[key] = (key, ko, en)
# 광역시/도의 한글명(정식 명칭)으로도 찾을 수 있게 등록
for key, ko, en in PROVINCE_TABLE:
    PROVINCE_LOOKUP[_norm_key(ko)] = (key, ko, en)

# 카카오 API가 실제로 주소에 쓰는 줄임말("서울특별시"가 아니라 "서울")도 등록.
# 이게 없으면 지오코딩 결과 주소로 city를 채우는 로직이 거의 항상 실패함
# (Kakao road_address_name/address_name은 정식 명칭이 아니라 이 줄임말을 씀).
_SHORT_PROVINCE_NAMES = {
    "seoul": "서울", "busan": "부산", "daegu": "대구", "incheon": "인천",
    "gwangju": "광주", "daejeon": "대전", "ulsan": "울산", "sejong": "세종",
    "gyeonggi": "경기", "gangwon": "강원",
    "chungcheongbuk": "충북", "chungcheongnam": "충남",
    "jeollabuk": "전북", "jeollanam": "전남",
    "gyeongsangbuk": "경북", "gyeongsangnam": "경남", "jeju": "제주",
}
for key, short_ko in _SHORT_PROVINCE_NAMES.items():
    _, ko, en = PROVINCE_LOOKUP[key]
    PROVINCE_LOOKUP[_norm_key(short_ko)] = (key, ko, en)


_ADMIN_SUFFIXES = ("teukbyeoljachisi", "teukbyeoljachido", "gwangyeoksi", "do", "si", "gun", "gu")


def resolve_province(token):
    """토큰 하나(영문 또는 한글)를 17개 광역시/도 중 하나로 매핑 시도.
    "Gyeongsangbuk-do", "Gyeongju-si"처럼 행정단위 접미사가 붙은 채로 오는 경우가
    많아서, 원본 키 그대로는 안 걸리고 접미사를 뗀 키로 재시도해야 함."""
    key = _norm_key(token)
    if not key:
        return None

    candidates = [key]
    for suf in _ADMIN_SUFFIXES:
        if key.endswith(suf) and len(key) > len(suf):
            candidates.append(key[: -len(suf)])

    for cand in candidates:
        if cand in PROVINCE_LOOKUP:
            return PROVINCE_LOOKUP[cand]
        if cand in DISTRICT_TO_PROVINCE and DISTRICT_TO_PROVINCE[cand]:
            return PROVINCE_LOOKUP[DISTRICT_TO_PROVINCE[cand]]
    return None


def parse_location(loc):
    """주소/지역 문자열 -> (city_id, city_name_ko, city_name_en, is_domestic, matched)."""
    if not loc or not str(loc).strip() or str(loc).strip() in {"-", "—"}:
        return None, None, None, None, False

    raw = str(loc).strip()

    # 해외 지명 체크
    for kw in NON_KOREA_KEYWORDS:
        if kw in raw.lower():
            return None, None, None, False, False

    # 쉼표로 분리, 뒤에서부터 매칭 시도 (상위 지역이 보통 마지막에 옴)
    parts = [p.strip() for p in raw.replace("—", "").split(",") if p.strip()]
    if not parts:
        # 쉼표 없는 한 덩어리 주소 (한글 주소 등) -> 공백 토큰으로 재시도
        parts = raw.split()

    for part in reversed(parts):
        # 한글 주소는 접미사 단위로 토큰을 다시 쪼갬 (예: "서울특별시 강남구 신사동...")
        sub_tokens = re.findall(
            r"[가-힣]{2,8}(?:특별시|광역시|특별자치시|특별자치도|도|시|군|구)", part
        )
        # 단어 단위 폴백은 짧은 조각(실제 주소 토큰)에만 적용.
        # 긴 문장(설명문 등)에 낱말 단위로 적용하면 "...represent Seoul."처럼
        # 본문 중 우연히 등장하는 지명 단어를 주소로 오인식하는 문제가 있었음.
        words = part.split()
        if len(words) <= 5:
            sub_tokens += words
        for tok in sub_tokens + [part]:
            match = resolve_province(tok)
            if match:
                city_id, city_ko, city_en = match
                return city_id, city_ko, city_en, True, tok

    return None, None, None, True, None  # 국내로 보이나 지역 특정 실패


# ============================================================
# 2. 카테고리 정규화
# ============================================================

CATEGORY_RULES = [
    (("restaurant", "bbq", "noodle", "tteokbokki", "bossam", "kalguksu", "hotpot",
      "bindaetteok", "gopchang", "pig house", "sushi"), "음식점"),
    (("cafe", "café", "bakery", "dessert", "bubble tea", "yogurt", "brunch", "tea"), "카페/디저트"),
    (("shopping", "store", "shop", "market", "merch", "arcade", "boutique"), "쇼핑"),
    (("filming", "mv "), "촬영지"),
    (("museum", "gallery", "exhibition"), "박물관/전시"),
    (("theme park",), "테마파크"),
    (("park",), "공원"),
    (("beach",), "해변"),
    (("palace", "historic", "museum village"), "유적지"),
    (("agency", " hq", "entertainment"), "기획사"),
    (("concert", "dome", "stadium", "hall", "arena", "venue"), "공연장"),
    (("accommodation", "resort", "hotel"), "숙박"),
    (("landmark", "tower", "observatory"), "랜드마크"),
    (("photo",), "포토스팟"),
    (("sauna", "jjimjilbang"), "체험/휴양"),
    (("ski", "cruise", "activity", "sports"), "체험/액티비티"),
]


def categorize(raw_category):
    if not raw_category:
        return None
    s = str(raw_category).lower()
    for keywords, mapped in CATEGORY_RULES:
        if any(kw in s for kw in keywords):
            return mapped
    return None


FOOD_CATEGORIES = {"음식점", "카페/디저트"}


# ============================================================
# 3. 이름 정규화 / 중복 병합 (김지윤 개발일지의 normalize_place 방식 참고)
# ============================================================

def normalize_place_key(name):
    if not name:
        return ""
    s = re.sub(r"\[[^\]]*\]", "", name)  # [운영중지] 같은 상태표시 제거
    s = re.sub(r"^\d+\.\s*", "", s)  # 목록 번호 제거
    s = re.sub(r"[\s\-_/·,()]", "", s)
    return s.lower().strip()


def is_same_place(a, b):
    """주의: 0.6 임계값이었을 때 "KQ Entertainment(ATEEZ 소속사)"와 "SM Entertainment
    Building"(Red Velvet 소속사)처럼 완전히 다른 기획사가 "Entertainment"라는
    공통 단어 때문에 0.61~0.64로 걸려서 같은 장소로 잘못 합쳐진 사례가 실제로
    발견됨. 0.82로 올려서 이런 "공통 단어 하나 때문에 유사해 보이는" 케이스를
    거르고, 진짜 표기 차이(오타/공백 수준)만 남도록 함."""
    ka, kb = normalize_place_key(a), normalize_place_key(b)
    if not ka or not kb:
        return False
    if ka == kb:
        return True
    if len(ka) >= 4 and len(kb) >= 4 and (ka in kb or kb in ka):
        return True
    return difflib.SequenceMatcher(None, ka, kb).ratio() >= 0.82


def slugify(name):
    if not name:
        return "unknown"
    s = re.sub(r"[^a-zA-Z0-9가-힣]+", "_", name).strip("_").lower()
    return s or "unknown"


def slugify_hyphen(name):
    """새 template(Place.id)용 - "전부 소문자·하이픈" 규칙. 한글 이름만 있는 경우엔
    로마자 변환기가 없어서 한글을 그대로 두고 공백/특수문자만 하이픈으로 바꿈
    (한글은 대소문자 개념이 없어 "소문자" 규칙과 충돌하지 않음)."""
    if not name:
        return "unknown"
    s = re.sub(r"[^a-zA-Z0-9가-힣]+", "-", name).strip("-").lower()
    return s or "unknown"


ARTIST_ALIASES = {
    # 2026-08-24 팀 결정: 그룹 활동명이 "(G)I-DLE"에서 "I-DLE"로 바뀌어서, 원본에
    # "(G)I-DLE"로 크롤링된 것도 최종 데이터셋엔 "I-DLE"로 통일해서 저장한다.
    "idle": "I-DLE", "gidle": "I-DLE", "i-dle": "I-DLE",
    "girlsgeneration": "Girls' Generation", "snsd": "Girls' Generation",
    "straykids": "Stray Kids", "skz": "Stray Kids",
    # NCT DREAM/NCT 127은 별도 그룹이 아니라 NCT 산하 유닛이라 상위 그룹(NCT)으로
    # 통일해서 수집한다 (2026-08-24 팀 결정).
    "nctdream": "NCT", "nct127": "NCT",
}


def canonical_artist(name):
    if not name:
        return None
    key = re.sub(r"[^a-z0-9]", "", name.lower())
    return ARTIST_ALIASES.get(key, name.strip())


def _artist_key(name):
    return re.sub(r"[^a-z0-9가-힣]", "", (name or "").lower())


def load_artist_allowlist():
    """artist_allowlist.json(구독자 수 기준으로 엄선된 29개 그룹 + 소속 멤버) 로딩.
    그룹명/멤버명 둘 다 허용 대상 -> place row의 artist_name이 그룹으로도,
    개별 멤버명으로도 들어올 수 있어서 둘 다 모아야 함."""
    path = os.path.join(BASE_DIR, "artist_allowlist.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    allowed = set()
    for entry in data:
        allowed.add(_artist_key(canonical_artist(entry["group"])))
        for member in entry["members"]:
            allowed.add(_artist_key(canonical_artist(member)))
    return allowed


def filter_manual_exclusions(rows):
    kept = []
    for r in rows:
        name = (r.get("place_name_en") or r.get("place_name_ko") or "").strip().lower()
        artist = r.get("artist_name")
        if (artist, name) in MANUAL_EXCLUSIONS:
            continue
        kept.append(r)
    return kept


def filter_by_allowlist(rows, allowed_keys):
    """리스트에 없는 아티스트(그룹/개인)가 붙은 place row는 제외.
    artist_name이 애초에 없는 행(특정 그룹에 귀속 안 됨)은 배제 대상이 아니라서 유지.
    비교 전에 canonical_artist로 별칭 정규화부터 함 ((G)I-DLE vs i-dle 같은 표기
    차이 때문에 원문 그대로 키 비교하면 같은 그룹인데도 다르게 잡히는 문제 있었음)."""
    kept, dropped_artists = [], set()
    for r in rows:
        artist = r.get("artist_name")
        if not artist or _artist_key(canonical_artist(artist)) in allowed_keys:
            kept.append(r)
        else:
            dropped_artists.add(artist)
    return kept, dropped_artists


NAME_MENTION_PATTERN = re.compile(r"([A-Z][A-Za-z.\-]{1,20})(?:\s*&\s*([A-Z][A-Za-z.\-]{1,20}))?'s\b")


def load_member_group_map():
    """artist_allowlist.json -> {정규화된 이름(그룹명 또는 멤버명): 그룹명}."""
    path = os.path.join(BASE_DIR, "artist_allowlist.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    mapping = {}
    for entry in data:
        group = entry["group"]
        mapping[_artist_key(canonical_artist(group))] = group
        for member in entry["members"]:
            mapping[_artist_key(canonical_artist(member))] = group
    return mapping


def extract_mentioned_names(text):
    """"Sehun & Xiumin's curry spot", "Jimin's father's restaurant" 같은 문장에서
    소유격('s) 앞의 인명 후보를 뽑음. "Sehun & Xiumin's"처럼 &로 묶인 것도 둘 다 잡음."""
    names = []
    for m in NAME_MENTION_PATTERN.finditer(text or ""):
        names.append(m.group(1))
        if m.group(2):
            names.append(m.group(2))
    return names


def filter_by_relation_consistency(rows, member_group_map):
    """artist_name과 relation_text_en에 소유격으로 언급된 인물이 서로 다른
    그룹이면 제외. (병합 버그로 엉뚱한 아티스트가 붙은 경우의 안전망 - 예:
    "Jungwon's Happy Graduation vlog" 텍스트가 artist_name=TWICE로 잘못 붙는 경우)
    언급된 이름이 allowlist에 아예 없는 사람(매니저, 다른 연예인 등)이면 판단
    근거가 없으니 그대로 둠 - 확신 있을 때만 제외."""
    kept, dropped = [], []
    for r in rows:
        artist = r.get("artist_name")
        text = r.get("relation_text_en") or ""
        if not artist or not text:
            kept.append(r)
            continue

        artist_group = member_group_map.get(_artist_key(canonical_artist(artist)))
        if not artist_group:
            kept.append(r)  # allowlist 밖 아티스트는 이미 앞 단계에서 걸러짐
            continue

        mentioned_groups = {
            member_group_map[_artist_key(name)]
            for name in extract_mentioned_names(text)
            if _artist_key(name) in member_group_map
        }
        if mentioned_groups and artist_group not in mentioned_groups:
            dropped.append((artist, text, sorted(mentioned_groups)))
        else:
            kept.append(r)
    return kept, dropped


def split_bilingual_name(name):
    """'Sushi Tatsu (스시타츠)' 같은 표기 -> (영문명, 한글명)."""
    if not name:
        return None, None
    m = re.match(r"^(.*?)\(([^)]+)\)\s*$", name.strip())
    if m:
        outer, inner = m.group(1).strip(), m.group(2).strip()
        if HANGUL_RE.search(inner) and not HANGUL_RE.search(outer):
            return (outer or None), inner
        if HANGUL_RE.search(outer) and not HANGUL_RE.search(inner):
            return inner, outer
    if HANGUL_RE.search(name):
        return None, name.strip()
    return name.strip(), None


# ============================================================
# 4. 소스별 로더
# ============================================================

def load_tour_spots():
    """Kpop_Tour_Spots_Combined.xlsx - 가장 깨끗한 구조화 소스."""
    path = os.path.join(RAW_DIR, "Kpop_Tour_Spots_Combined.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["All K-pop Tour Spots"]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        artist = r[0]
        place = r[1] if len(r) > 1 else None
        cat = r[2] if len(r) > 2 else None
        loc = r[3] if len(r) > 3 else None
        notes = r[4] if len(r) > 4 else None
        src = r[5] if len(r) > 5 else None
        if not place:
            continue

        place_en, place_ko = split_bilingual_name(str(place).strip())
        city_id, city_ko, city_en, is_domestic, _ = parse_location(loc)
        mapped_cat = categorize(cat)

        review_reasons = []
        if city_id is None:
            review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님(해외 촬영/방문지 추정)")
        if mapped_cat is None:
            review_reasons.append(f"카테고리 미분류(원본: {cat!r})")

        rows.append({
            "artist_name": canonical_artist(str(artist).strip()),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "place_name_ko": place_ko, "place_name_en": place_en,
            "address": str(loc).strip() if loc else None,
            "relation_text_en": str(notes).strip() if notes else None,
            "place_category": mapped_cat,
            "source_url": str(src).strip() if src else None,
            "is_food": mapped_cat in FOOD_CATEGORIES if mapped_cat else None,
            "is_local_spot": None,
            "status": "needs_review" if review_reasons else "ready",
            "_review_reason": "; ".join(review_reasons),
            "_source": "Kpop_Tour_Spots_Combined.xlsx",
        })
    return rows


def load_thisismykorea():
    path = os.path.join(RAW_DIR, "stara_thisismykorea.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    for item in data:
        place = (item.get("place") or "").strip()
        if not place:
            continue

        # 진짜 장소명이면 보통 12단어, 마침표 1개를 안 넘음. 그보다 길거나 문장이
        # 여러 개면 place/description 필드가 밀려서 설명문이 들어온 오염 행임
        # (ATEEZ 3건 확인, 2026-08-24 - desc 필드엔 "Address :" 라벨만 남고 진짜
        # 장소명은 원본 어디에도 없어서 자동 복구 불가). 살릴 수 없으니 통째로 폐기.
        if len(place.split()) > 12 or place.count(".") >= 2:
            continue

        review_reasons = []
        place_en, place_ko = split_bilingual_name(place)

        addr = item.get("address")
        # 실제 주소라면 번지수(숫자)가 거의 항상 포함됨. 숫자 없이 8단어 넘는 긴 문자열은
        # place/description/address 필드가 밀려 들어간 오염 행일 가능성이 높아 주소로 신뢰하지 않음.
        addr_text = str(addr or "")
        looks_like_address = bool(re.search(r"\d", addr_text)) or len(addr_text.split()) <= 8
        if not looks_like_address:
            city_id = city_ko = city_en = None
            review_reasons.append("주소 형식 아님(필드 밀림 의심, 원본 데이터 확인 필요)")
        else:
            city_id, city_ko, city_en, is_domestic, _ = parse_location(addr)
            if city_id is None:
                review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님")

        rows.append({
            "artist_name": canonical_artist((item.get("artist") or "").strip() or None),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "place_name_ko": place_ko, "place_name_en": place_en,
            "address": addr if looks_like_address else None,
            "relation_text_en": (item.get("description") or "").strip() or None,
            "source_url": item.get("url"),
            "is_food": None,
            "is_local_spot": None,
            "status": "needs_review" if review_reasons else "ready",
            "_review_reason": "; ".join(review_reasons),
            "_source": "stara_thisismykorea.json",
        })
    return rows


# 사람이 직접 검토하고 "이건 빼라"고 확정한 (아티스트, 장소명) 쌍.
# 예: "Chinatown" 검색은 지역 전체를 가리키는 이름이라 카카오에 단일 POI로
# 없고, 매번 그 이름이 들어간 딴 업체(주차장/게스트하우스 등)에 잘못 매칭됨 ->
# 자동 매칭 로직으로 못 고치는 케이스라 수동 제외.
MANUAL_EXCLUSIONS = {
    ("EXO", "chinatown"),
}


# "장소가 아닌 목차/질문/안내성 문구" 판별 - README가 요구한 "미분류/노이즈 필터"에 해당
NOISE_START_WORDS = (
    "why ", "how ", "is ", "are ", "where ", "what ", "understanding ",
    "best ", "sample ", "a full-day", "tax-free", "online ", "stadium / dome tour",
)


def clean_listicle_title(name):
    """'2. KWANGYA@SEOUL & SM Town COEX (Seoul Forest / Gangnam)' 같은 표기 정리.
    앞 번호 제거, ' — 부제'/' - 부제'는 관계 텍스트로 분리."""
    s = re.sub(r"^\d+\.\s*", "", name.strip())
    subtitle = None
    m = re.match(r"^(.*?)\s+[—–]\s+(.*)$", s)
    if m:
        s, subtitle = m.group(1).strip(), m.group(2).strip()
    return s, subtitle


def is_noise_title(name):
    s = name.strip().lower()
    if s.endswith("?"):
        return True
    return any(s.startswith(w) for w in NOISE_START_WORDS)


def artist_mentioned_in_text(artist_raw, text):
    """아티스트명이 place/description 원문에 실제로 등장하는지 확인(대소문자 무시).
    stara_places.json을 직접 까보니 73건 중 31건(43%)에 "IVE" 태그가 붙어있는데,
    뷰티숍/드라마 촬영지/일반 공연장 소개처럼 IVE와 무관한 항목에도 죄다 붙어있는
    스크래핑 오염이 발견됨(2026-08-24 확인) - 본문에 이름이 실제로 언급된
    아티스트만 신뢰하는 걸로 걸러냄. IVE 하나만의 문제가 아니라 이 소스의
    artists 필드 전체를 못 믿는다는 뜻이라 일반 규칙으로 적용."""
    pattern = r"(?<![A-Za-z0-9])" + re.escape(artist_raw) + r"(?![A-Za-z0-9])"
    return re.search(pattern, text, re.IGNORECASE) is not None


# stara_places.json 전용 - 노이즈 제목 필터/아티스트 언급 검증을 통과하고도 여전히
# "특정 성지 하나"로 볼 수 없는 항목들을 사람이 직접 확인해서 제외 목록으로 확정함
# (2026-08-24). 사유는 각 줄에 표기: 장소 여러 개를 한 행에 묶은 "블록", 목차성
# 절 제목, 특정 시즌 이벤트 근거(이 정적 데이터셋의 기존 제외 기준에 해당).
STARA_PLACES_DROP_TITLES = {
    "kwangya@seoul & sm town coex (seoul forest / gangnam)",  # 장소 2곳 묶임
    "cube café & cube studio (seongsu / gangnam)",  # 뒤에 나오는 단일 "Cube Cafe"와 중복 + 장소 묶임
    "photocard & boba cafés (hongdae, sinchon)",  # 특정 업체가 아니라 여러 지역 카테고리
    "understanding the korean wave (hallyu)",  # 주제 개괄문, 장소 아님
    "official label stores",  # 여러 기획사 매장을 한 행에 묶음
    "stadium / dome tour",  # 절 제목, 장소 아님
    "national museum of korea — 2026 blackpink partnership",  # 2026-02 한정 특별전 - 시즌 이벤트라 기존 제외 기준(README "장소 포함/제외 기준")에 해당
}


def load_stara_places():
    path = os.path.join(RAW_DIR, "stara_places.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    dropped_noise = 0
    dropped_no_artist = 0
    for item in data:
        raw_place = (item.get("place") or "").strip()
        if not raw_place:
            continue
        cleaned, subtitle = clean_listicle_title(raw_place)
        if is_noise_title(cleaned) or is_noise_title(raw_place):
            dropped_noise += 1
            continue
        # 제외 목록은 부제(em-dash 뒷부분)까지 포함한 원제목 기준으로 등록했으므로,
        # clean_listicle_title()이 분리하기 전의 번호만 뗀 제목으로 대조해야 한다.
        unsplit_title = re.sub(r"^\d+\.\s*", "", raw_place).strip().lower()
        if unsplit_title in STARA_PLACES_DROP_TITLES:
            dropped_noise += 1
            continue

        base_desc = (item.get("description") or "").strip()
        text = f"{raw_place} {base_desc}"
        artists = [a for a in (item.get("artists") or []) if artist_mentioned_in_text(a, text)]
        if not artists:
            # 태그가 아예 없었거나(장소 소개/구 가이드류), 있어도 본문에 근거가
            # 없는 오태깅(위 IVE 사례)뿐 - 성지로 볼 근거가 없어 폐기.
            dropped_no_artist += 1
            continue

        place_en, place_ko = split_bilingual_name(cleaned)

        # 이 소스는 주소/지역 필드가 아예 없음 -> city는 항상 비어있음.
        # (URL이 seoultourism.org 계열이라 실제로는 대부분 서울이겠지만, 문서 단위로
        # 확정된 정보가 아니라서 추측 대입하지 않고 지오코딩/사람 확인으로 넘김)
        review_reasons = ["city 정보 없음(소스에 주소 필드 없음)"]

        relation = base_desc
        if subtitle:
            relation = (subtitle + (" - " + base_desc if base_desc else "")).strip()

        for artist in artists:
            rows.append({
                "artist_name": canonical_artist(artist),
                "place_name_ko": place_ko, "place_name_en": place_en,
                "relation_text_en": relation or None,
                "image_url": item.get("image") or None,
                "source_url": item.get("url"),
                "is_food": None, "is_local_spot": None,
                "status": "needs_review",
                "_review_reason": "; ".join(review_reasons),
                "_source": "stara_places.json",
            })
    print(f"  stara_places.json: {len(data)}건 중 노이즈/블록 {dropped_noise}건, "
          f"아티스트 근거 없음 {dropped_no_artist}건 제외")
    return rows


def load_manual_places():
    """raw_data_정인지/*.json - 팀원이 직접 블로그 등에서 확인하고 정리한 장소.
    이미 정확한 주소/카테고리를 사람이 확인해서 넣은 소스라 다른 로더보다
    신뢰도가 높음 (category_hint는 우리 taxonomy로 이미 분류된 값이라
    categorize() 안 거치고 그대로 씀). 새 블로그를 추가하려면 이 폴더에
    같은 스키마의 json 파일을 추가하면 자동으로 인식됨."""
    import glob
    rows = []
    for path in glob.glob(os.path.join(BASE_DIR, "raw_data_정인지", "*.json")):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            place_ko = (item.get("place_name_ko") or "").strip()
            if not place_ko:
                continue
            city_id, city_ko, city_en, is_domestic, _ = parse_location(item.get("address"))

            review_reasons = []
            if item.get("needs_manual_check"):
                review_reasons.append("사람이 직접 표시한 확인 필요 항목(원본 데이터 내 주소 불일치 등)")
            if city_id is None:
                review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님")

            cat = item.get("category_hint")
            rows.append({
                # artist_name은 그룹 단위로 일관되게 유지(allowlist/GROUP_KO_NAMES가
                # 그룹 기준이라서) - 멤버는 아래 _manual_member로 따로 전달.
                "artist_name": canonical_artist(item.get("artist")),
                "_manual_member": (item.get("member") or "").strip() or None,
                # ↑ relation_text 문장에서 정규식으로 멤버를 추출하는
                # resolve_artist_ids()의 기본 방식은 "OO's 인물"처럼 소유격이
                # 인물명 앞에 와야 잡히는데, 이 소스는 "BTS's Jin"처럼 그룹명에
                # 소유격이 붙고 멤버명이 뒤에 오는 문장이 많아서 정규식이 못 잡음
                # (2026-08-24 인천 재검수 중 발견 - 원본에 이미 있는 명시적
                # member 필드를 그동안 안 읽고 있었음). 있으면 이 값을 우선 사용.
                "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
                "place_name_ko": place_ko, "place_name_en": None,
                "address": item.get("address"),
                "relation_text_ko": item.get("description_ko"),
                "relation_text_en": item.get("description_en"),
                "place_category": cat,
                "source_url": item.get("url"),
                "is_food": cat in FOOD_CATEGORIES if cat else None,
                "is_local_spot": None,
                "status": "needs_review" if review_reasons else "ready",
                "_review_reason": "; ".join(review_reasons),
                "_source": os.path.basename(path),
            })
    return rows


def load_hometowns():
    """곽채원 xlsx + 정인지 CSV 합집합. (Idol, Group) 기준 중복 제거,
    누락 필드는 서로 보완."""
    combined = {}

    xlsx_path = os.path.join(RAW_DIR, "Kpop_Idol_Hometowns_filtered.xlsx")
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb["Idol Hometowns"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            idol, group, province, district = r[0], r[1], r[2], r[3]
            key = (str(idol).strip().lower(), str(group).strip().lower())
            combined[key] = {"idol": idol, "group": group, "province": province, "district": district}

    import glob
    csv_candidates = glob.glob(os.path.join(BASE_DIR, "아이돌_출신지_*.csv"))
    if not csv_candidates:
        csv_candidates = glob.glob(os.path.join(BASE_DIR, "raw_data_정인지", "아이돌_출신지_*.csv"))
    if not csv_candidates:
        raise FileNotFoundError("아이돌_출신지_*.csv 파일을 찾을 수 없음")
    csv_path = csv_candidates[0]
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # 빈 첫 줄
        header = next(reader)  # Idol,Group,Province,...
        for r in reader:
            if not r or not r[0]:
                continue
            idol, group, province, district = r[0], r[1], r[2], r[3]
            key = (str(idol).strip().lower(), str(group).strip().lower())
            existing = combined.get(key, {})
            combined[key] = {
                "idol": idol,
                "group": group,
                "province": province or existing.get("province"),
                "district": district or existing.get("district"),
            }

    rows = []
    for rec in combined.values():
        location = rec["district"] or rec["province"]
        city_id, city_ko, city_en, is_domestic, _ = parse_location(location)
        if city_id is None and rec["province"]:
            city_id, city_ko, city_en, is_domestic, _ = parse_location(rec["province"])

        rows.append({
            "artist_name": canonical_artist(str(rec["idol"]).strip()),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "relation_text_ko": f"{rec['idol']}의 고향",
            "relation_text_en": f"{rec['idol']}'s hometown",
            "quest_type": "hometown",
            "is_food": False,
            "is_local_spot": True,
            "status": "needs_review" if city_id is None else "ready",
            "_review_reason": "" if city_id is not None else "도시 매핑 실패",
            "_source": "hometowns(xlsx+csv 병합)",
            "_group": rec["group"],
        })
    return rows


# ============================================================
# 5. 병합(중복 제거) - place 계열만 대상. hometown은 애초에 (Idol,Group) 단위라 제외.
# ============================================================

def merge_places(rows):
    """같은 장소로 판단되는 행들을 묶되, relation_text/source_url처럼 "이 아티스트와
    이 장소의 관계"를 설명하는 필드는 아티스트별로 따로 유지함.
    예전엔 그룹의 대표 행(group[0]) 하나의 relation_text를 그룹 내 모든 아티스트에게
    그대로 복사해서 붙였는데, 실제로 같은 장소를 여러 아티스트가 각자 다른 이유로
    방문한 정당한 케이스(예: 파라다이스시티 - ENHYPEN 소속사 행사 vs BTS 뷔 앰버서더)에서
    서로 다른 relation_text가 뒤섞이는 문제가 있었음. place 자체의 공통 정보
    (이름/카테고리/이미지)만 그룹 전체에서 공유하고, 관계 설명은 아티스트 단위로 분리."""
    merged = []
    used = [False] * len(rows)
    for i, row in enumerate(rows):
        if used[i]:
            continue
        group = [row]
        used[i] = True
        name_i = row.get("place_name_en") or row.get("place_name_ko")
        for j in range(i + 1, len(rows)):
            if used[j]:
                continue
            other = rows[j]
            name_j = other.get("place_name_en") or other.get("place_name_ko")
            if row.get("city_id") == other.get("city_id") and is_same_place(name_i, name_j):
                group.append(other)
                used[j] = True

        shared = {}
        for field in ("place_name_ko", "place_name_en", "place_category", "image_url", "address"):
            for g in group:
                if g.get(field):
                    shared[field] = g[field]
                    break

        by_artist = OrderedDict()
        for g in group:
            by_artist.setdefault(g.get("artist_name"), []).append(g)

        for artist, artist_rows in by_artist.items():
            base = artist_rows[0]
            sources = sorted({g["source_url"] for g in artist_rows if g.get("source_url")})
            relations_en = sorted({g["relation_text_en"] for g in artist_rows if g.get("relation_text_en")})
            relations_ko = sorted({g["relation_text_ko"] for g in artist_rows if g.get("relation_text_ko")})
            reason_fragments = set()
            for g in artist_rows:
                for frag in (g.get("_review_reason") or "").split("; "):
                    if frag:
                        reason_fragments.add(frag)
            # "카테고리 미분류" 사유는 place 단위 속성이라, 그룹 내 다른 아티스트의
            # 행에서 카테고리가 이미 채워졌다면(shared) 이 아티스트에게도 더 이상
            # 유효한 사유가 아님 - 제거.
            if shared.get("place_category"):
                reason_fragments = {f for f in reason_fragments if not f.startswith("카테고리 미분류")}
            reasons = sorted(reason_fragments)

            merged_row = dict(base)
            merged_row.update(shared)
            merged_row["artist_name"] = artist
            merged_row["source_url"] = "; ".join(sources) if sources else base.get("source_url")
            merged_row["relation_text_en"] = " / ".join(relations_en) if relations_en else base.get("relation_text_en")
            merged_row["relation_text_ko"] = " / ".join(relations_ko) if relations_ko else base.get("relation_text_ko")
            merged_row["_review_reason"] = "; ".join(reasons)
            merged_row["status"] = "needs_review" if reasons else "ready"
            merged.append(merged_row)
    return merged

CACHE_PATH = os.path.join(OUT_DIR, ".geocode_cache.json")


def load_env():
    """STARA/.env를 읽어서 os.environ에 반영 (python-dotenv 의존성 없이 최소 구현)."""
    env_path = os.path.join(BASE_DIR, ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip()
            if key and key not in os.environ:
                os.environ[key] = val


def load_geocode_cache():
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_geocode_cache(cache):
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def kakao_keyword_search(query, api_key):
    resp = requests.get(
        "https://dapi.kakao.com/v2/local/search/keyword.json",
        params={"query": query, "size": 5},
        headers={"Authorization": f"KakaoAK {api_key}"},
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json().get("documents", [])


# 관광/방문 목적지가 될 수 없는 카테고리 - 이런 매칭은 이름만 비슷하고 실제로는
# 전혀 다른 시설인 경우가 많음 (예: "Chinatown" 검색이 "차이나타운 공영주차장"에
# 매칭된 사례, "Citizen Park" 검색이 근처 아파트 단지에 매칭된 사례,
# "UN Village"(백현 노래 제목이자 동네 이름) 검색이 "유엔빌리지치과의원"에
# 매칭된 사례가 실제로 확인됨(2026-08-25) - K-pop 성지는 절대 이런 업종일 수
# 없으니 무조건 제외. 사람이 매번 "이거 진짜 같은 곳 맞아?"를 눈으로 확인해야
# 했던 이유 중 상당수가 사실 이 필터 하나로 걸러지는 기계적 오매칭이었음 -
# 이 목록에 안 걸리는 나머지(예: "아차산"을 검색했는데 이름이 비슷한 실내
# 클라이밍짐 지점에 매칭되는 경우처럼 카테고리 자체는 그럴듯한 케이스)는
# 카테고리만으로 못 걸러서 여전히 "영문명으로 검색한 매칭" 리뷰가 필요함.
NOISE_CATEGORIES = (
    "주차장", "부동산", "공인중개사", "아파트", "주거시설", "오피스텔",
    "병원", "의원", "치과", "한의원", "약국", "동물병원",
    "학원", "교습소", "어린이집", "유치원",
    "은행", "금융", "보험",
    "장례식장", "묘지", "납골당",
    "관공서", "주민센터", "구청", "시청", "경찰서", "소방서", "세무서", "법원", "검찰청",
    "정비소", "세차장", "주유소",
)


def pick_best_match(docs, city_ko):
    """도시명이 결과 주소에 포함된 것만 신뢰 (엉뚱한 지역의 동명 장소 오매칭 방지).
    주의: 이것만으로는 부족함 - "Busan Citizen Park" 검색이 "시민공원아파트"(아파트
    단지)에 매칭되는 것처럼, 도시만 맞고 실제로는 다른 장소인 오탐이 확인됨.
    그래서 이 결과를 쓰는 쪽(geocode_and_fill_names)에서 영문 일반명사성 쿼리는
    무조건 needs_review로 내려서 사람이 좌표를 눈으로 확인하게 함."""
    docs = [d for d in docs if not any(nc in (d.get("category_name") or "") for nc in NOISE_CATEGORIES)]
    if not docs:
        return None
    if not city_ko:
        return docs[0]
    city_key = city_ko.replace("특별시", "").replace("광역시", "").replace("특별자치시", "").replace("특별자치도", "").replace("도", "")
    for doc in docs:
        addr = (doc.get("road_address_name") or "") + (doc.get("address_name") or "")
        if city_key and city_key in addr:
            return doc
    return None  # 결과는 있지만 다 다른 도시 -> 신뢰 안 함


# Kakao category_name(예: "음식점 > 한식 > 육류,고기")을 우리 place_category taxonomy로 매핑
KAKAO_CATEGORY_RULES = [
    (("음식점",), "음식점"),
    (("카페", "제과,베이커리", "디저트"), "카페/디저트"),
    (("쇼핑", "판매", "마트", "백화점", "패션"), "쇼핑"),
    (("박물관", "전시관", "미술관"), "박물관/전시"),
    (("테마파크",), "테마파크"),
    (("공원",), "공원"),
    (("해수욕장", "해변"), "해변"),
    (("고궁", "유적", "사적"), "유적지"),
    (("공연장", "콘서트홀", "경기장", "운동장"), "공연장"),
    (("숙박", "호텔", "모텔", "펜션", "리조트"), "숙박"),
    (("관광,명소", "랜드마크", "타워"), "랜드마크"),
    (("사진", "포토"), "포토스팟"),
    (("찜질방", "사우나"), "체험/휴양"),
    (("스포츠", "레저", "액티비티", "낚시"), "체험/액티비티"),
]


def categorize_kakao(category_name):
    if not category_name:
        return None
    for keywords, mapped in KAKAO_CATEGORY_RULES:
        if any(kw in category_name for kw in keywords):
            return mapped
    return None


def geocode_and_fill_names(rows, api_key):
    cache = load_geocode_cache()
    n_queried = n_hit = n_cached = n_skipped = 0

    # city_id 유무는 대상 조건에 없음 - city 미상 행(주소 필드 자체가 없는 소스 등)도
    # 장소명만 있으면 대상에 포함시켜서, 아래에서 카카오 검색 결과 주소로 city를
    # 역으로 채워볼 기회를 준다(2026-08-24, "city 미상 67건" 보강 시도 결정).
    targets = [
        r for r in rows
        if (r.get("place_name_ko") or r.get("place_name_en"))
        and (r.get("latitude") is None or r.get("place_name_ko") is None)
    ]
    print(f"   지오코딩 대상: {len(targets)}행 (전체 {len(rows)}행 중 장소명 있는 것)")

    for i, row in enumerate(targets, 1):
        had_korean_name = bool(row.get("place_name_ko"))
        name = row.get("place_name_ko") or row.get("place_name_en")
        city_en = row.get("city_name_en") or ""
        query = f"{city_en} {name}".strip()

        if query in cache:
            docs = cache[query]
            n_cached += 1
        else:
            try:
                docs = kakao_keyword_search(query, api_key)
            except requests.RequestException as e:
                # Windows 콘솔(cp949)이 원본 텍스트의 특수문자를 못 그리는 경우가 있어서
                # 에러 로그 출력 자체가 죽지 않게 ascii로 안전하게 치환해서 출력.
                safe_query = query.encode("ascii", "replace").decode("ascii")
                print(f"   [실패] {safe_query!r} -> {e}")
                docs = []
            cache[query] = docs
            n_queried += 1
            time.sleep(0.15)  # 초당 요청 제한 여유

        if i % 50 == 0:
            print(f"   ...{i}/{len(targets)} 처리, 신규 조회 {n_queried}건")

        best = pick_best_match(docs, row.get("city_name_ko"))
        if not best:
            frag = "카카오 키워드 검색 매칭 없음(이름/좌표 확인 필요)"
            existing = row.get("_review_reason") or ""
            if frag not in existing:
                row["_review_reason"] = (existing + "; " + frag).strip("; ")
            row["status"] = "needs_review"
            continue

        n_hit += 1
        if not row.get("place_name_ko"):
            row["place_name_ko"] = best.get("place_name")
        row["latitude"] = float(best["y"])
        row["longitude"] = float(best["x"])

        # city가 원래 없었으면(주소 필드 자체가 없던 stara_places.json 등),
        # 카카오가 돌려준 실제 주소를 우리 parse_location으로 파싱해서 채움.
        if not row.get("city_id"):
            addr = best.get("road_address_name") or best.get("address_name")
            city_id, city_ko, city_en, _, _ = parse_location(addr)
            if city_id:
                row["city_id"], row["city_name_ko"], row["city_name_en"] = city_id, city_ko, city_en
                fragments = [f for f in (row.get("_review_reason") or "").split("; ")
                             if f and not f.startswith("도시 매핑 실패")
                             and not f.startswith("city 정보 없음")
                             and not f.startswith("국내 지명 아님")]
                row["_review_reason"] = "; ".join(fragments)

        # 카테고리가 이걸로 새로 채워졌으면, 로딩 단계에서 붙었던
        # "카테고리 미분류(...)" 사유는 더 이상 유효하지 않으니 제거.
        if not row.get("place_category"):
            mapped_cat = categorize_kakao(best.get("category_name"))
            if mapped_cat:
                row["place_category"] = mapped_cat
                row["is_food"] = mapped_cat in FOOD_CATEGORIES
                fragments = [f for f in (row.get("_review_reason") or "").split("; ")
                             if f and not f.startswith("카테고리 미분류")]
                row["_review_reason"] = "; ".join(fragments)

        if not had_korean_name:
            # 원래 영문(설명형) 이름으로 검색한 매칭은 "도시가 같다"만 확인됐을 뿐 실제
            # 동일 장소인지는 검증이 약함(예: "Citizen Park" -> 근처 아파트단지가
            # 걸린 사례 실제 확인됨). 좌표는 채워두되 사람이 한 번 봐야 함.
            frag = "영문명으로 검색한 매칭(실제 동일 장소인지 확인 필요)"
            existing = row.get("_review_reason") or ""
            if frag not in existing:
                row["_review_reason"] = (existing + "; " + frag).strip("; ")

        # 남은 사유가 없으면 ready로 승격, 있으면 needs_review 유지.
        row["status"] = "needs_review" if row.get("_review_reason") else "ready"

    save_geocode_cache(cache)
    print(f"   완료: 신규조회 {n_queried} / 캐시재사용 {n_cached} / 매칭성공 {n_hit} / 대상 {len(targets)}")




# ============================================================
# 8. 최종 조립 - cities.json / artists.json / places.json
# ============================================================
#
# 앞으로 계속 이 3파일 구조로 운영하기로 해서, place_id/artist_id 부여·멤버 단위
# 분리·출처 URL 검증을 전부 이 단계에서 자동으로 처리함 (예전엔 인천 데이터를
# 손으로 만들면서 사람이 직접 판단했던 부분들). merge_places()가 만든
# (장소,아티스트) 단위 행을 다시 장소 단위로 묶어서 최종 스키마로 변환한다.
#
# 주의: 아래 두 한글명 테이블(GROUP_KO_NAMES/MEMBER_KO_NAMES)에 없는
# 아티스트는 한글명 자리에 영문명을 임시로 넣고 needs_ko_name으로 표시함 -
# 틀린 한글명을 지어내는 것보다 "아직 모른다"고 표시하는 게 안전함.

# 2026-08-20 팀 회의에서 카테고리 체계를 food/shopping/culture/activity/
# landmark_observatory/kpop 6종으로 확정 (기존 photo/experience 폐지).
# "숙박"은 의도적으로 여기 없음 - 숙소류는 카테고리가 아니라 스코프 자체에서
# 제외 대상이라, build_schema()에서 이 행 자체를 드롭한다(ACCOMMODATION_LABELS 참고).
CATEGORY_TO_TS = {
    "음식점": "food", "카페/디저트": "food",
    "쇼핑": "shopping",
    "박물관/전시": "culture", "유적지": "culture", "공연장": "culture",
    "기획사": "kpop",  # 소속사 사옥(하이브 인사이트 등)은 culture가 아니라 kpop
    "공원": "activity", "해변": "activity", "테마파크": "activity",
    "체험/액티비티": "activity", "체험/휴양": "activity",
    # 순수 포토존은 별도 카테고리 없이 landmark_observatory로 흡수(회의 결정)
    "랜드마크": "landmark_observatory", "포토스팟": "landmark_observatory",
    "촬영지": "landmark_observatory",
}

VALID_PLACE_CATEGORIES = {"food", "shopping", "culture", "activity", "landmark_observatory", "kpop"}

# raw_data_정인지/*.json의 category_hint처럼 이미 최종 영문 카테고리값이 들어오는
# 소스도 있어서, 그 경우는 번역 없이 그대로 통과시킨다(둘 다 안 걸리면 activity로 폴백 -
# 예전 "experience" 캐치올 자리를 대신함).
def resolve_category(cat_ko):
    if cat_ko in VALID_PLACE_CATEGORIES:
        return cat_ko
    return CATEGORY_TO_TS.get(cat_ko, "activity")


ACCOMMODATION_LABELS = {"숙박", "accommodation", "resort", "hotel", "hostel", "guesthouse", "펜션", "풀빌라", "게스트하우스"}


GROUP_KO_NAMES = {
    "bigbang": "빅뱅", "girlsgeneration": "소녀시대", "2ne1": "투애니원",
    "exo": "엑소", "bts": "방탄소년단", "mamamoo": "마마무", "redvelvet": "레드벨벳",
    "winner": "위너", "ikon": "아이콘", "seventeen": "세븐틴", "straykids": "스트레이 키즈",
    "twice": "트와이스", "astro": "아스트로", "nct": "엔시티", "ateez": "에이티즈",
    "idle": "(여자)아이들", "itzy": "있지", "txt": "투모로우바이투게더", "aespa": "에스파",
    "enhypen": "엔하이픈", "treasure": "트레저", "ive": "아이브", "lesserafim": "르세라핌",
    "newjeans": "뉴진스", "nmixx": "엔믹스", "babymonster": "베이비몬스터",
    "illit": "아일릿", "cortis": "코르티스", "blackpink": "블랙핑크",
}

# (그룹_artist_key, 멤버_artist_key) -> 한글 이름. 확실히 아는 것만.
MEMBER_KO_NAMES = {
    ("bts", "v"): "뷔", ("bts", "jin"): "진", ("bts", "jimin"): "지민", ("bts", "rm"): "알엠",
    ("enhypen", "sunoo"): "선우", ("enhypen", "jungwon"): "정원", ("enhypen", "sunghoon"): "성훈",
    ("twice", "tzuyu"): "쯔위", ("twice", "chaeyoung"): "채영", ("twice", "dahyun"): "다현",
    ("twice", "jihyo"): "지효", ("twice", "nayeon"): "나연",
    ("exo", "baekhyun"): "백현", ("exo", "chanyeol"): "찬열", ("exo", "chen"): "첸",
    ("exo", "do"): "디오", ("exo", "kai"): "카이", ("exo", "xiumin"): "시우민",
    ("nct", "taeyong"): "태용",
    ("redvelvet", "joy"): "조이",
    ("straykids", "in"): "아이엔",
    ("txt", "beomgyu"): "범규", ("txt", "soobin"): "수빈", ("txt", "yeonjun"): "연준",
    ("bigbang", "gdragon"): "지드래곤", ("bigbang", "taeyang"): "태양",
}

DWELL_MINUTES_BY_CATEGORY = {"food": 50, "landmark_observatory": 25, "shopping": 50}
# culture/activity/kpop은 spec에 평균 기준이 없어서 의도적으로 비움 (None)


def looks_like_url(s):
    return bool(s) and bool(re.match(r"^https?://", s.strip()))


def validate_source_url(raw):
    """"Trazy", "User-provided list" 같은 텍스트 라벨은 실제 링크가 아니므로
    거르고, 진짜 URL만 남김. 하나도 없으면 "PENDING"으로 통일해서 나중에
    grep으로 쉽게 찾을 수 있게 함."""
    if not raw:
        return "PENDING"
    urls = [u.strip() for u in raw.split(";") if looks_like_url(u.strip())]
    return "; ".join(dict.fromkeys(urls)) if urls else "PENDING"


# 2026-08-25 정책 변경: 케이팝 팬덤 특성상 링크 하나 없이도 작은 단서(멤버가
# 언급한 콘텐츠 한 조각)만으로 실제 촬영지/방문지를 특정해내는 경우가 많다.
# source_url이 진짜 링크가 아니어도, relation_text 자체가 "누가 + 어떤
# 콘텐츠에서" 나왔는지 구체적으로 밝히고 있으면(나중에 실제로 찾아볼 수 있는
# 추적 가능한 인용이면) 출처 미상으로 보지 않는다. 다만 이 기준을 통과해도
# status는 draft 그대로 유지 - "확인 가능한 링크"로 사람이 직접 검증해야
# verified로 올라간다는 원칙은 안 바뀜(review 표시만 해제).
# 판별 기준은 "인물 + 구체적 콘텐츠명"(날짜는 필수 아님) - "Jimin's father's
# restaurant"처럼 근거가 될 콘텐츠 자체가 없는 맨주장은 여전히 불인정.
# 이 목록/휴리스틱은 완벽하지 않으므로(새 콘텐츠 포맷이 계속 생김), 애매한
# 케이스는 사람이 직접 판단하는 게 안전함.
CREDIBLE_CONTENT_MARKERS = (
    "vlog", "-log", "weverse", "instagram", "twitter", "x post", "youtube",
    "music video", "episode", "cover", "livestream", "concert", "fancam",
    "broadcast", "interview", "behind", "pop-up", "guide", "sev sev tour",
    "diary", "telepathy", "code ep", "1n2d", "so so fun", "run bts",
    "racha log", "skz code",
)
# 위 목록은 문장 아무 데나 있어도 됨(부분 문자열 검색). 아래는 그 자체로는
# 너무 짧고 흔해서(다른 단어에 우연히 포함될 수 있어서 - "big"/"deep" 등)
# 단어 경계로만 매칭한다. "Vernon, ... Vernon IG"처럼 문장 맨 끝에 오면
# 뒤에 공백/쉼표가 없어서 예전엔 안 걸렸던 버그를 여기서 고침(2026-08-25,
# 부산 조현화랑 항목 재검증 중 발견).
CREDIBLE_CONTENT_WORD_MARKERS = ("ig", "mv", "ep")
HEDGE_WORDS = ("reportedly", "not exact", "unconfirmed", "알려진", "추정")


def has_credible_citation(relation_text):
    if not relation_text:
        return False
    text = relation_text.lower()
    if any(h in text for h in HEDGE_WORDS):
        return False
    if any(m in text for m in CREDIBLE_CONTENT_MARKERS):
        return True
    return any(re.search(r"\b" + re.escape(w) + r"\b", text) for w in CREDIBLE_CONTENT_WORD_MARKERS)


def resolve_artist_ids(artist_rows, member_group_map):
    """이 장소를 공유하는 (아티스트, 관계텍스트) 행들을 보고, 특정 멤버 개인
    얘기면 "{그룹}-{멤버}"로 분리하고, 아니면 그룹 그대로 둠.
    각 행 "자신의" relation_text만 보고 판단함 - 합쳐진 텍스트를 보면 다른
    아티스트 얘기까지 섞여서 오탐 위험이 있었음(이전 버그 참고).
    한계: "Sehun & Xiumin's" 같은 소유격 패턴만 잡음 - "TWICE's Tzuyu,
    Chaeyoung, and Dahyun"처럼 그룹명에 소유격이 붙고 멤버가 나열되는
    문장은 못 잡아서 그룹 단위로 남음 (사람이 검토해서 보완 필요).
    반환: [(artist_id, group_name, member_name_or_None), ...] 순서 유지, 중복 제거."""
    results = []
    seen = set()
    for ar in artist_rows:
        group = ar.get("artist_name")
        if not group:
            continue
        group_key = _artist_key(canonical_artist(group))

        # raw_data_정인지처럼 원본에 명시적 member 필드가 있으면 이걸 최우선으로
        # 신뢰한다 - 정규식 추출보다 훨씬 정확함("BTS's Jin"처럼 그룹명에
        # 소유격이 붙는 문장은 정규식으로 못 잡는데, 이런 소스가 실제로 있었음).
        manual_member = ar.get("_manual_member")
        if manual_member:
            mentioned = [manual_member]
        else:
            text = " ".join(filter(None, [ar.get("relation_text_ko"), ar.get("relation_text_en")]))
            mentioned = []
            for name in extract_mentioned_names(text):
                key = _artist_key(name)
                g = member_group_map.get(key)
                if g and _artist_key(g) == group_key and key != group_key:
                    mentioned.append(name)

        if mentioned:
            for m in dict.fromkeys(mentioned):
                aid = f"{slugify_hyphen(group)}-{slugify_hyphen(m)}"
                if aid not in seen:
                    seen.add(aid)
                    results.append((aid, group, m))
        else:
            aid = slugify_hyphen(group)
            if aid not in seen:
                seen.add(aid)
                results.append((aid, group, None))
    return results


def build_schema(place_rows, member_group_map):
    """merge_places() + geocode 이후의 (장소,아티스트) 단위 행을 장소 단위로
    다시 묶어서 cities/artists/places 최종 스키마로 조립.
    좌표가 없는 행은 required 필드를 못 채우므로 places.json에서 빼고
    별도로 보고한다(dropped_no_coords)."""
    by_place = OrderedDict()
    for r in place_rows:
        key = (r.get("city_id"), r.get("place_name_ko") or "", r.get("place_name_en") or "")
        by_place.setdefault(key, []).append(r)

    places = []
    city_registry = {}
    artist_registry = {}
    dropped_no_coords = []
    dropped_accommodation = []
    dropped_no_city = []

    for (city_id, name_ko, name_en), group in by_place.items():
        if not city_id:
            # --geocode를 쓰면 geocode_and_fill_names()가 카카오 검색 결과 주소로
            # city를 역으로 채우는 걸 이미 시도한 뒤임(2026-08-24부터 city 미상
            # 행도 지오코딩 대상에 포함되도록 함) - 그래도 안 채워진 행만 여기로
            # 옴. 그냥 버리지 않고 개수/목록을 남겨서 사람이 원본을 볼 수 있게 함.
            dropped_no_city.append({
                "place_name_ko": name_ko, "place_name_en": name_en,
                "artists": sorted({g.get("artist_name") for g in group if g.get("artist_name")}),
            })
            continue

        lat = next((g.get("latitude") for g in group if g.get("latitude") is not None), None)
        lon = next((g.get("longitude") for g in group if g.get("longitude") is not None), None)
        if lat is None or lon is None:
            dropped_no_coords.append({
                "city_id": city_id, "place_name_ko": name_ko, "place_name_en": name_en,
                "artists": sorted({g.get("artist_name") for g in group if g.get("artist_name")}),
            })
            continue

        cat_ko = next((g.get("place_category") for g in group if g.get("place_category")), None)
        if isinstance(cat_ko, str) and cat_ko.strip().lower() in ACCOMMODATION_LABELS:
            # 숙소류(호텔/리조트/펜션/풀빌라)는 카테고리가 아니라 스코프 자체에서
            # 제외 - 좌표/영업시간이 확인돼도 상관없이 뺀다(2026-08-20 회의 결정)
            dropped_accommodation.append({
                "city_id": city_id, "place_name_ko": name_ko, "place_name_en": name_en,
                "artists": sorted({g.get("artist_name") for g in group if g.get("artist_name")}),
            })
            continue

        needs_name_review = not (name_ko and name_en)
        final_name_ko = name_ko or name_en
        final_name_en = name_en or name_ko

        cat_en = resolve_category(cat_ko)

        resolved = resolve_artist_ids(group, member_group_map)
        artist_ids = [aid for aid, _, _ in resolved]
        for aid, group_name, member_name in resolved:
            group_key = _artist_key(canonical_artist(group_name))
            group_ko = GROUP_KO_NAMES.get(group_key)
            if member_name:
                member_ko = MEMBER_KO_NAMES.get((group_key, _artist_key(member_name)))
                artist_registry.setdefault(aid, [member_ko or member_name, member_name, member_ko is None])
            else:
                artist_registry.setdefault(aid, [group_ko or group_name, group_name, group_ko is None])

        relations_ko = list(dict.fromkeys(g["relation_text_ko"] for g in group if g.get("relation_text_ko")))
        relations_en = list(dict.fromkeys(g["relation_text_en"] for g in group if g.get("relation_text_en")))
        sources = list(dict.fromkeys(g["source_url"] for g in group if g.get("source_url")))
        source_url = validate_source_url("; ".join(sources))
        # 실제 링크가 없어도 relation_text 자체가 추적 가능한 인용이면
        # "출처 미상"으로 안 본다(2026-08-25 정책, has_credible_citation 참고).
        if source_url == "PENDING":
            combined_relation = " / ".join(relations_en) or " / ".join(relations_ko)
            if has_credible_citation(combined_relation):
                source_url = f"[콘텐츠 인용 - 링크 미확인] {combined_relation}"

        place_slug = slugify_hyphen(final_name_en or final_name_ko)
        place_id = f"{artist_ids[0]}-{cat_en}-{place_slug}" if artist_ids else f"unknown-{cat_en}-{place_slug}"

        image_url = next((g.get("image_url") for g in group if g.get("image_url")), None)
        address = next((g.get("address") for g in group if g.get("address")), None)

        review_notes = []
        if needs_name_review:
            review_notes.append("한 언어 이름 없음(다른 언어로 임시 대체)")
        if source_url == "PENDING":
            review_notes.append("실제 출처 URL 없음")
        if not relations_ko:
            review_notes.append("한글 relation_text 없음")
        if not relations_en:
            review_notes.append("영문 relation_text 없음")
        # 지오코딩 단계(geocode_and_fill_names)가 붙여둔 사유(영문명으로 검색한
        # 매칭이라 오탐 위험/카카오 매칭 자체 실패 등)를 여기서 매번 새로
        # 만드느라 놓치고 있었음 - 좌표 정확도가 최우선 순위인 프로젝트에서
        # 가장 위험한 신호(오탐 가능성)가 최종 review CSV에 하나도 안 뜨던
        # 버그였음(2026-08-24 서울 검수 중 발견 - 356건 중 129건이 이 사유였음).
        for g in group:
            for frag in (g.get("_review_reason") or "").split("; "):
                if frag and frag not in review_notes:
                    review_notes.append(frag)

        places.append({
            "place_id": place_id,
            "city_id": city_id,
            "artist_ids": artist_ids,
            "place_name_ko": final_name_ko,
            "place_name_en": final_name_en,
            "place_category": cat_en,
            "latitude": round(lat, 6),
            "longitude": round(lon, 6),
            "relation_text_ko": " / ".join(relations_ko) if relations_ko else "[한글 설명 필요]",
            "relation_text_en": " / ".join(relations_en) if relations_en else "[EN description needed]",
            "source_url": source_url,
            "open_time": None,
            "close_time": None,
            "dwell_minutes": DWELL_MINUTES_BY_CATEGORY.get(cat_en),
            "quest_type": None,
            "quest_text_ko": None,
            "quest_text_en": None,
            "image_url": image_url,
            "tour_api_content_id": None,
            "is_food": cat_en == "food",
            "is_local_spot": False,
            "status": "draft",
            "_address": address,
            "_review_notes": "; ".join(review_notes),
        })

        city_ko = next((g.get("city_name_ko") for g in group if g.get("city_name_ko")), None)
        city_en = next((g.get("city_name_en") for g in group if g.get("city_name_en")), None)
        city_registry[city_id] = (city_ko, city_en)

    return places, city_registry, artist_registry, dropped_no_coords, dropped_accommodation, dropped_no_city


def generated_filename(kind, ext, city_filter, owner_filter):
    """산출물 파일명 규칙(2026-08-25 확정) - "generated_{종류}_{도시}_{작성자}.{확장자}"
    형태로, 종류가 맨 앞·도시/작성자는 뒤쪽·구분자는 전부 언더바.
    도시/작성자가 없으면(--city=/--owner= 생략) 그 부분은 그냥 빠진다.
    예: generated_places_seoul_정인지.json, generated_places.json(둘 다 생략 시)."""
    suffix_parts = [p for p in (city_filter, owner_filter) if p]
    suffix = ("_" + "_".join(suffix_parts)) if suffix_parts else ""
    return f"generated_{kind}{suffix}.{ext}"


def save_schema(places, city_registry, artist_registry, city_filter, owner_filter):
    os.makedirs(OUT_DIR, exist_ok=True)

    cities = []
    for city_id, (city_ko, city_en) in sorted(city_registry.items()):
        short_ko = _SHORT_PROVINCE_NAMES.get(city_id, city_ko)
        cities.append({"city_id": city_id, "city_name_ko": short_ko, "city_name_en": city_en})

    artists = []
    needs_ko_name = []
    for artist_id, (ko, en, needs_ko) in sorted(artist_registry.items()):
        artists.append({"artist_id": artist_id, "artist_name_ko": ko, "artist_name_en": en})
        if needs_ko:
            needs_ko_name.append(artist_id)

    places_out = [{k: v for k, v in p.items() if not k.startswith("_")} for p in places]

    cities_path = os.path.join(OUT_DIR, generated_filename("cities", "json", city_filter, owner_filter))
    artists_path = os.path.join(OUT_DIR, generated_filename("artists", "json", city_filter, owner_filter))
    places_path = os.path.join(OUT_DIR, generated_filename("places", "json", city_filter, owner_filter))
    review_path = os.path.join(OUT_DIR, generated_filename("places_needs_review", "csv", city_filter, owner_filter))

    with open(cities_path, "w", encoding="utf-8") as f:
        json.dump(cities, f, ensure_ascii=False, indent=2)
    with open(artists_path, "w", encoding="utf-8") as f:
        json.dump(artists, f, ensure_ascii=False, indent=2)
    with open(places_path, "w", encoding="utf-8") as f:
        json.dump(places_out, f, ensure_ascii=False, indent=2)

    review_places = [p for p in places if p.get("_review_notes")]
    with open(review_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["place_id", "city_id", "artist_ids", "place_name_ko", "_review_notes"])
        writer.writeheader()
        for p in review_places:
            writer.writerow({
                "place_id": p["place_id"], "city_id": p["city_id"],
                "artist_ids": ";".join(p["artist_ids"]), "place_name_ko": p["place_name_ko"],
                "_review_notes": p["_review_notes"],
            })

    return {
        "cities": len(cities), "artists": len(artists), "places": len(places),
        "needs_ko_name": needs_ko_name, "needs_review": len(review_places),
        "cities_path": cities_path, "artists_path": artists_path,
        "places_path": places_path, "review_path": review_path,
    }


def main():
    print("1) Kpop_Tour_Spots_Combined.xlsx 로딩...")
    tour_rows = load_tour_spots()
    print(f"   {len(tour_rows)}행")

    print("2) stara_thisismykorea.json 로딩...")
    tmk_rows = load_thisismykorea()
    print(f"   {len(tmk_rows)}행")

    print("3) stara_places.json 로딩 (노이즈 필터 적용)...")
    sp_rows = load_stara_places()
    print(f"   {len(sp_rows)}행")

    print("3b) raw_data_정인지/*.json 로딩 (직접 확인한 블로그 등)...")
    manual_rows = load_manual_places()
    print(f"   {len(manual_rows)}행")

    print("4) 아이돌 출신지 (xlsx + CSV) 로딩...")
    home_rows = load_hometowns()
    print(f"   {len(home_rows)}행 - 참고용으로만 집계함. 현재 Place 스키마는 특정 위경도가 "
          f"있는 '장소'만 다루도록 되어 있어서, 출신지(도시 단위) 데이터를 어떻게 편입할지는 "
          f"아직 스키마 상 정해진 게 없음 -> 이번 출력에는 포함하지 않음")

    print("5) 장소 데이터(1~3b) 중복 병합...")
    place_rows = merge_places(tour_rows + tmk_rows + sp_rows + manual_rows)
    print(f"   병합 전 {len(tour_rows) + len(tmk_rows) + len(sp_rows) + len(manual_rows)}행 -> 병합 후 {len(place_rows)}행")

    place_rows = filter_manual_exclusions(place_rows)

    print("5b) 아티스트 allowlist(artist_allowlist.json) 필터링...")
    allowed_keys = load_artist_allowlist()
    place_rows, dropped_artists = filter_by_allowlist(place_rows, allowed_keys)
    print(f"   리스트 밖 아티스트 {len(dropped_artists)}종 제외: {sorted(dropped_artists)}")
    print(f"   {len(place_rows)}행 남음")

    print("5c) artist_name vs relation_text 인물 불일치 필터링...")
    member_group_map = load_member_group_map()
    place_rows, dropped_mismatches = filter_by_relation_consistency(place_rows, member_group_map)
    print(f"   불일치로 {len(dropped_mismatches)}행 제외")
    print(f"   {len(place_rows)}행 남음")

    if "--geocode" in sys.argv:
        print("6) 지오코딩 + 한글 상호명 채우기 (Kakao 키워드 검색)...")
        load_env()
        api_key = os.environ.get("KAKAO_REST_API_KEY", "").strip()
        if not api_key:
            print("   [건너뜀] KAKAO_REST_API_KEY가 .env에 없음")
        else:
            geocode_and_fill_names(place_rows, api_key)
            # merge_places()(5단계)는 city_id가 아직 안 채워진 상태에서 돌기
            # 때문에, 주소 필드 자체가 없는 소스(stara_places.json 등) 유래
            # 행은 다른 소스에 이미 있는 같은 실제 장소와 병합되지 못하고
            # 남아있다가 여기서 각자 다른 지오코딩 결과를 받는 경우가 있었음
            # (2026-08-24 서울 검수 중 발견 - "서울숲"이 진짜 서울숲과 5.6km
            # 떨어진 "홍릉시험림"으로 잘못 매칭된 채 별개 행으로 남아있었음).
            # city_id가 다 채워진 지금 다시 한 번 병합해서 이런 케이스를 합친다.
            print("6b) 지오코딩으로 city 채워진 뒤 중복 병합 재시도...")
            before_n = len(place_rows)
            place_rows = merge_places(place_rows)
            print(f"   {before_n}행 -> {len(place_rows)}행")
    else:
        print("6) 지오코딩 건너뜀 (실행하려면: python build_dataset.py --geocode)")

    city_filter = None
    owner_filter = None
    for arg in sys.argv:
        if arg.startswith("--city="):
            city_filter = arg.split("=", 1)[1].strip()
        elif arg.startswith("--owner="):
            owner_filter = arg.split("=", 1)[1].strip()
    if city_filter:
        place_rows = [r for r in place_rows if r.get("city_id") == city_filter]
        print(f"   --city={city_filter} 필터 적용: {len(place_rows)}행")

    print("7) 최종 조립 - cities/artists/places.json...")
    places, city_registry, artist_registry, dropped_no_coords, dropped_accommodation, dropped_no_city = build_schema(place_rows, member_group_map)
    # 자동 생성 결과는 항상 .generated.json으로 따로 저장하고, 사람이 손으로
    # 검수/보완한 cities.json/artists.json/places.json은 절대 자동으로 덮어쓰지
    # 않는다 - 한 번 이 스크립트가 빈 결과로 정본 파일을 덮어써서 복구한 적이
    # 있어서(지오코딩 없이 돌리면 좌표가 없어 전부 걸러짐), 재발 방지.
    # 검수 후 정본으로 승격하려면 사람이 직접 파일을 바꿔치기할 것.
    stats = save_schema(places, city_registry, artist_registry, city_filter, owner_filter)
    print(f"   cities: {stats['cities']} / artists: {stats['artists']} / places: {stats['places']}")
    print(f"   (자동 생성본 - 정본 cities.json/artists.json/places.json은 건드리지 않음. "
          f"검수 후 필요하면 사람이 직접 교체할 것)")
    print(f"   좌표 없어서 제외된 곳: {len(dropped_no_coords)}건")
    print(f"   숙소류라서 제외된 곳: {len(dropped_accommodation)}건")
    print(f"   지오코딩 후에도 city 미상이라 제외된 곳: {len(dropped_no_city)}건")
    if dropped_no_city:
        dropped_no_city_path = os.path.join(
            OUT_DIR, generated_filename("dropped_no_city", "csv", city_filter, owner_filter))
        with open(dropped_no_city_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["place_name_ko", "place_name_en", "artists"])
            writer.writeheader()
            for d in dropped_no_city:
                writer.writerow({
                    "place_name_ko": d["place_name_ko"], "place_name_en": d["place_name_en"],
                    "artists": ";".join(d["artists"]),
                })
        print(f"   -> {dropped_no_city_path} 에 목록 저장 (원본 재확인용)")
    print(f"   review 필요: {stats['needs_review']}건 -> {stats['review_path']}")
    print(f"   한글 아티스트명 미확인(임시로 영문명 사용): {len(stats['needs_ko_name'])}종 {stats['needs_ko_name']}")
    print(f"-> {stats['cities_path']}, {stats['artists_path']}, {stats['places_path']}")


if __name__ == "__main__":
    main()
