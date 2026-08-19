# -*- coding: utf-8 -*-
"""
raw_data_곽채원/ + 아이돌_출신지_정인지 CSV -> Data_Preprocessing_Template.json 스키마로 정제/병합.

입력 소스:
    1. Kpop_Tour_Spots_Combined.xlsx ("All K-pop Tour Spots" 시트) - 462행, 가장 깨끗한 구조화 데이터
    2. stara_thisismykorea.json - 23행, 구조화됨 (visitkorea.or.kr)
    3. stara_places.json - 73행, 기사 목차형 텍스트라 노이즈 필터링 필요
    4. Kpop_Idol_Hometowns_filtered.xlsx + 아이돌_출신지_정인지 CSV - 아이돌 출신지(고향), 두 소스 합집합

출력:
    preprocessed/dataset.csv, preprocessed/dataset.json - 전체 결과 (template 스키마)
    preprocessed/needs_review.csv - status=needs_review 행만, review_reason 컬럼 추가

주의:
    - 위도/경도는 지오코딩 API 키가 없어서 비워둠 (다음 단계)
    - city_name은 광역시/도 17개 단위로 통일 (구/시 단위 세부 주소는 데이터 소스마다
      한글/영문이 섞여있어서, 항상 안정적으로 뽑히는 상위 단위로 맞춤)
    - 확신 없는 행은 지우지 않고 status=needs_review로 남겨서 사람이 검수하게 함
"""

import csv
import json
import os
import re
import sys
import time
import difflib
from collections import OrderedDict

import openpyxl
import requests

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "raw_data_곽채원")
OUT_DIR = os.path.join(BASE_DIR, "preprocessed")

TEMPLATE_FIELDS = [
    "artist_id", "artist_name", "city_id", "city_name_ko", "city_name_en",
    "place_id", "place_name_ko", "place_name_en", "relation_text_ko", "relation_text_en",
    "latitude", "longitude", "open_time", "close_time", "dwell_minutes",
    "place_category", "quest_type", "quest_text_ko", "quest_text_en",
    "image_url", "source_url", "tour_api_content_id",
    "is_food", "is_local_spot", "status",
]

HANGUL_RE = re.compile(r"[가-힣]")

# ============================================================
# 1. 지역(광역시/도) 매핑 - 17개 단위로 통일
# ============================================================

# 정규화 키(소문자, 공백/하이픈 제거) -> (city_id, city_name_ko, city_name_en)
PROVINCE_TABLE = [
    ("seoul", "서울특별시", "Seoul"),
    ("busan", "부산광역시", "Busan"),
    ("daegu", "대구광역시", "Daegu"),
    ("incheon", "인천광역시", "Incheon"),
    ("gwangju", "광주광역시", "Gwangju"),
    ("daejeon", "대전광역시", "Daejeon"),
    ("ulsan", "울산광역시", "Ulsan"),
    ("sejong", "세종특별자치시", "Sejong"),
    ("gyeonggi", "경기도", "Gyeonggi"),
    ("gangwon", "강원특별자치도", "Gangwon"),
    ("chungcheongbuk", "충청북도", "Chungcheongbuk"),
    ("chungbuk", "충청북도", "Chungcheongbuk"),
    ("chungcheongnam", "충청남도", "Chungcheongnam"),
    ("chungnam", "충청남도", "Chungcheongnam"),
    ("jeollabuk", "전라북도", "Jeollabuk"),
    ("jeonbuk", "전라북도", "Jeollabuk"),
    ("jeollanam", "전라남도", "Jeollanam"),
    ("jeonnam", "전라남도", "Jeollanam"),
    ("gyeongsangbuk", "경상북도", "Gyeongsangbuk"),
    ("gyeongbuk", "경상북도", "Gyeongsangbuk"),
    ("gyeongsangnam", "경상남도", "Gyeongsangnam"),
    ("gyeongnam", "경상남도", "Gyeongsangnam"),
    ("jeju", "제주특별자치도", "Jeju"),
]

# 구/시/군 이름 -> 소속 광역시/도 정규화 키 (영문 주소에서 상위 지역명이
# 아예 안 나오고 구/시 이름만 있는 경우 보강용. 데이터에 실제 등장하는 이름만 등록)
DISTRICT_TO_PROVINCE = {
    # Seoul 25개 구
    "jung": "seoul", "jongno": "seoul", "yongsan": "seoul", "seongdong": "seoul",
    "gwangjin": "seoul", "dongdaemun": "seoul", "jungnang": "seoul", "seongbuk": "seoul",
    "gangbuk": "seoul", "dobong": "seoul", "nowon": "seoul", "eunpyeong": "seoul",
    "seodaemun": "seoul", "mapo": "seoul", "yangcheon": "seoul", "gangseo": "seoul",
    "guro": "seoul", "geumcheon": "seoul", "yeongdeungpo": "seoul", "dongjak": "seoul",
    "gwanak": "seoul", "seocho": "seoul", "gangnam": "seoul", "songpa": "seoul",
    "gangdong": "seoul",
    "myeongdong": "seoul", "hongdae": "seoul", "apgujeong": "seoul", "cheongdam": "seoul",
    "itaewon": "seoul", "insadong": "seoul", "hannam": "seoul", "seongsu": "seoul",
    "hanam": "seoul",  # xlsx 소스에 "Hanam, Seoul"로 등장 (구 단위 취급됨, 원자료 그대로)
    # Busan
    "haeundae": "busan", "gijang": "busan", "dongnae": "busan", "suyeong": "busan",
    "sasang": "busan", "saha": "busan",
    # Gyeonggi 주요 시
    "goyang": "gyeonggi", "gwacheon": "gyeonggi", "bucheon": "gyeonggi", "guri": "gyeonggi",
    "siheung": "gyeonggi", "ansan": "gyeonggi", "namyangju": "gyeonggi", "yongin": "gyeonggi",
    "anyang": "gyeonggi", "hwaseong": "gyeonggi", "gunpo": "gyeonggi", "seongnam": "gyeonggi",
    "suwon": "gyeonggi", "gimpo": "gyeonggi", "uijeongbu": "gyeonggi",
    # 기타 (경상/전라/강원/제주 등 시 단위)
    "suncheon": "jeollanam", "changwon": "gyeongsangnam", "sinan": "jeollanam",
    "chungju": "chungcheongbuk", "iksan": "jeollabuk", "jeonju": "jeollabuk",
    "yangsan": "gyeongsangnam", "pohang": "gyeongsangbuk", "namhae": "gyeongsangnam",
    "sacheon": "gyeongsangnam", "chuncheon": "gangwon", "gangneung": "gangwon",
    "gyeongju": "gyeongsangbuk", "jejucity": "jeju", "dalseo": "daegu",
    "bupyeong": "incheon", "namdong": "incheon", "geumjeong": "busan", "buk": None,  # 여러 도시에 중복 -> 단독 판단 불가
}

NON_KOREA_KEYWORDS = {
    "china", "japan", "uae", "dubai", "usa", "thailand", "hokkaido", "dongyang",
    "vietnam", "singapore", "taiwan", "hong kong", "malaysia", "philippines",
}


def _norm_key(s):
    # 영문은 소문자 알파벳만, 한글은 완성형 음절만 남김.
    # (주의: 원래 [^a-z]만 걸렀더니 한글 문자열은 전부 지워져서 빈 문자열이 되고,
    #  결과적으로 한글 주소는 단 하나도 지역 매칭이 안 되는 버그가 있었음)
    return re.sub(r"[^a-z가-힣]", "", s.lower())


PROVINCE_LOOKUP = {}
for key, ko, en in PROVINCE_TABLE:
    PROVINCE_LOOKUP[key] = (key, ko, en)
# 광역시/도의 한글명(정식 명칭)으로도 찾을 수 있게 등록
for key, ko, en in PROVINCE_TABLE:
    PROVINCE_LOOKUP[_norm_key(ko)] = (key, ko, en)

# 카카오 API가 실제로 주소에 쓰는 줄임말("서울특별시"가 아니라 "서울")도 등록.
# 이게 없으면 지오코딩 결과 주소로 city를 채우는 로직이 거의 항상 실패함
# (Kakao road_address_name/address_name은 정식 명칭이 아니라 이 줄임말을 씀).
_SHORT_PROVINCE_NAMES = {
    "seoul": "서울", "busan": "부산", "daegu": "대구", "incheon": "인천",
    "gwangju": "광주", "daejeon": "대전", "ulsan": "울산", "sejong": "세종",
    "gyeonggi": "경기", "gangwon": "강원",
    "chungcheongbuk": "충북", "chungcheongnam": "충남",
    "jeollabuk": "전북", "jeollanam": "전남",
    "gyeongsangbuk": "경북", "gyeongsangnam": "경남", "jeju": "제주",
}
for key, short_ko in _SHORT_PROVINCE_NAMES.items():
    _, ko, en = PROVINCE_LOOKUP[key]
    PROVINCE_LOOKUP[_norm_key(short_ko)] = (key, ko, en)


_ADMIN_SUFFIXES = ("teukbyeoljachisi", "teukbyeoljachido", "gwangyeoksi", "do", "si", "gun", "gu")


def resolve_province(token):
    """토큰 하나(영문 또는 한글)를 17개 광역시/도 중 하나로 매핑 시도.
    "Gyeongsangbuk-do", "Gyeongju-si"처럼 행정단위 접미사가 붙은 채로 오는 경우가
    많아서, 원본 키 그대로는 안 걸리고 접미사를 뗀 키로 재시도해야 함."""
    key = _norm_key(token)
    if not key:
        return None

    candidates = [key]
    for suf in _ADMIN_SUFFIXES:
        if key.endswith(suf) and len(key) > len(suf):
            candidates.append(key[: -len(suf)])

    for cand in candidates:
        if cand in PROVINCE_LOOKUP:
            return PROVINCE_LOOKUP[cand]
        if cand in DISTRICT_TO_PROVINCE and DISTRICT_TO_PROVINCE[cand]:
            return PROVINCE_LOOKUP[DISTRICT_TO_PROVINCE[cand]]
    return None


def parse_location(loc):
    """주소/지역 문자열 -> (city_id, city_name_ko, city_name_en, is_domestic, matched)."""
    if not loc or not str(loc).strip() or str(loc).strip() in {"-", "—"}:
        return None, None, None, None, False

    raw = str(loc).strip()

    # 해외 지명 체크
    for kw in NON_KOREA_KEYWORDS:
        if kw in raw.lower():
            return None, None, None, False, False

    # 쉼표로 분리, 뒤에서부터 매칭 시도 (상위 지역이 보통 마지막에 옴)
    parts = [p.strip() for p in raw.replace("—", "").split(",") if p.strip()]
    if not parts:
        # 쉼표 없는 한 덩어리 주소 (한글 주소 등) -> 공백 토큰으로 재시도
        parts = raw.split()

    for part in reversed(parts):
        # 한글 주소는 접미사 단위로 토큰을 다시 쪼갬 (예: "서울특별시 강남구 신사동...")
        sub_tokens = re.findall(
            r"[가-힣]{2,8}(?:특별시|광역시|특별자치시|특별자치도|도|시|군|구)", part
        )
        # 단어 단위 폴백은 짧은 조각(실제 주소 토큰)에만 적용.
        # 긴 문장(설명문 등)에 낱말 단위로 적용하면 "...represent Seoul."처럼
        # 본문 중 우연히 등장하는 지명 단어를 주소로 오인식하는 문제가 있었음.
        words = part.split()
        if len(words) <= 5:
            sub_tokens += words
        for tok in sub_tokens + [part]:
            match = resolve_province(tok)
            if match:
                city_id, city_ko, city_en = match
                return city_id, city_ko, city_en, True, tok

    return None, None, None, True, None  # 국내로 보이나 지역 특정 실패


# ============================================================
# 2. 카테고리 정규화
# ============================================================

CATEGORY_RULES = [
    (("restaurant", "bbq", "noodle", "tteokbokki", "bossam", "kalguksu", "hotpot",
      "bindaetteok", "gopchang", "pig house", "sushi"), "음식점"),
    (("cafe", "café", "bakery", "dessert", "bubble tea", "yogurt", "brunch", "tea"), "카페/디저트"),
    (("shopping", "store", "shop", "market", "merch", "arcade", "boutique"), "쇼핑"),
    (("filming", "mv "), "촬영지"),
    (("museum", "gallery", "exhibition"), "박물관/전시"),
    (("theme park",), "테마파크"),
    (("park",), "공원"),
    (("beach",), "해변"),
    (("palace", "historic", "museum village"), "유적지"),
    (("agency", " hq", "entertainment"), "기획사"),
    (("concert", "dome", "stadium", "hall", "arena", "venue"), "공연장"),
    (("accommodation", "resort", "hotel"), "숙박"),
    (("landmark", "tower", "observatory"), "랜드마크"),
    (("photo",), "포토스팟"),
    (("sauna", "jjimjilbang"), "체험/휴양"),
    (("ski", "cruise", "activity", "sports"), "체험/액티비티"),
]


def categorize(raw_category):
    if not raw_category:
        return None
    s = str(raw_category).lower()
    for keywords, mapped in CATEGORY_RULES:
        if any(kw in s for kw in keywords):
            return mapped
    return None


FOOD_CATEGORIES = {"음식점", "카페/디저트"}


# ============================================================
# 3. 이름 정규화 / 중복 병합 (김지윤 개발일지의 normalize_place 방식 참고)
# ============================================================

def normalize_place_key(name):
    if not name:
        return ""
    s = re.sub(r"\[[^\]]*\]", "", name)  # [운영중지] 같은 상태표시 제거
    s = re.sub(r"^\d+\.\s*", "", s)  # 목록 번호 제거
    s = re.sub(r"[\s\-_/·,()]", "", s)
    return s.lower().strip()


def is_same_place(a, b):
    """주의: 0.6 임계값이었을 때 "KQ Entertainment(ATEEZ 소속사)"와 "SM Entertainment
    Building"(Red Velvet 소속사)처럼 완전히 다른 기획사가 "Entertainment"라는
    공통 단어 때문에 0.61~0.64로 걸려서 같은 장소로 잘못 합쳐진 사례가 실제로
    발견됨. 0.82로 올려서 이런 "공통 단어 하나 때문에 유사해 보이는" 케이스를
    거르고, 진짜 표기 차이(오타/공백 수준)만 남도록 함."""
    ka, kb = normalize_place_key(a), normalize_place_key(b)
    if not ka or not kb:
        return False
    if ka == kb:
        return True
    if len(ka) >= 4 and len(kb) >= 4 and (ka in kb or kb in ka):
        return True
    return difflib.SequenceMatcher(None, ka, kb).ratio() >= 0.82


def slugify(name):
    if not name:
        return "unknown"
    s = re.sub(r"[^a-zA-Z0-9가-힣]+", "_", name).strip("_").lower()
    return s or "unknown"


def slugify_hyphen(name):
    """새 template(Place.id)용 - "전부 소문자·하이픈" 규칙. 한글 이름만 있는 경우엔
    로마자 변환기가 없어서 한글을 그대로 두고 공백/특수문자만 하이픈으로 바꿈
    (한글은 대소문자 개념이 없어 "소문자" 규칙과 충돌하지 않음)."""
    if not name:
        return "unknown"
    s = re.sub(r"[^a-zA-Z0-9가-힣]+", "-", name).strip("-").lower()
    return s or "unknown"


ARTIST_ALIASES = {
    "idle": "(G)I-DLE", "gidle": "(G)I-DLE", "i-dle": "(G)I-DLE",
    "girlsgeneration": "Girls' Generation", "snsd": "Girls' Generation",
    "straykids": "Stray Kids", "skz": "Stray Kids",
}


def canonical_artist(name):
    if not name:
        return None
    key = re.sub(r"[^a-z0-9]", "", name.lower())
    return ARTIST_ALIASES.get(key, name.strip())


def _artist_key(name):
    return re.sub(r"[^a-z0-9가-힣]", "", (name or "").lower())


def load_artist_allowlist():
    """artist_allowlist.json(구독자 수 기준으로 엄선된 29개 그룹 + 소속 멤버) 로딩.
    그룹명/멤버명 둘 다 허용 대상 -> place row의 artist_name이 그룹으로도,
    개별 멤버명으로도 들어올 수 있어서 둘 다 모아야 함."""
    path = os.path.join(BASE_DIR, "artist_allowlist.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    allowed = set()
    for entry in data:
        allowed.add(_artist_key(canonical_artist(entry["group"])))
        for member in entry["members"]:
            allowed.add(_artist_key(canonical_artist(member)))
    return allowed


def filter_manual_exclusions(rows):
    kept = []
    for r in rows:
        name = (r.get("place_name_en") or r.get("place_name_ko") or "").strip().lower()
        artist = r.get("artist_name")
        if (artist, name) in MANUAL_EXCLUSIONS:
            continue
        kept.append(r)
    return kept


def filter_by_allowlist(rows, allowed_keys):
    """리스트에 없는 아티스트(그룹/개인)가 붙은 place row는 제외.
    artist_name이 애초에 없는 행(특정 그룹에 귀속 안 됨)은 배제 대상이 아니라서 유지.
    비교 전에 canonical_artist로 별칭 정규화부터 함 ((G)I-DLE vs i-dle 같은 표기
    차이 때문에 원문 그대로 키 비교하면 같은 그룹인데도 다르게 잡히는 문제 있었음)."""
    kept, dropped_artists = [], set()
    for r in rows:
        artist = r.get("artist_name")
        if not artist or _artist_key(canonical_artist(artist)) in allowed_keys:
            kept.append(r)
        else:
            dropped_artists.add(artist)
    return kept, dropped_artists


NAME_MENTION_PATTERN = re.compile(r"([A-Z][A-Za-z.\-]{1,20})(?:\s*&\s*([A-Z][A-Za-z.\-]{1,20}))?'s\b")


def load_member_group_map():
    """artist_allowlist.json -> {정규화된 이름(그룹명 또는 멤버명): 그룹명}."""
    path = os.path.join(BASE_DIR, "artist_allowlist.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    mapping = {}
    for entry in data:
        group = entry["group"]
        mapping[_artist_key(canonical_artist(group))] = group
        for member in entry["members"]:
            mapping[_artist_key(canonical_artist(member))] = group
    return mapping


def extract_mentioned_names(text):
    """"Sehun & Xiumin's curry spot", "Jimin's father's restaurant" 같은 문장에서
    소유격('s) 앞의 인명 후보를 뽑음. "Sehun & Xiumin's"처럼 &로 묶인 것도 둘 다 잡음."""
    names = []
    for m in NAME_MENTION_PATTERN.finditer(text or ""):
        names.append(m.group(1))
        if m.group(2):
            names.append(m.group(2))
    return names


def filter_by_relation_consistency(rows, member_group_map):
    """artist_name과 relation_text_en에 소유격으로 언급된 인물이 서로 다른
    그룹이면 제외. (병합 버그로 엉뚱한 아티스트가 붙은 경우의 안전망 - 예:
    "Jungwon's Happy Graduation vlog" 텍스트가 artist_name=TWICE로 잘못 붙는 경우)
    언급된 이름이 allowlist에 아예 없는 사람(매니저, 다른 연예인 등)이면 판단
    근거가 없으니 그대로 둠 - 확신 있을 때만 제외."""
    kept, dropped = [], []
    for r in rows:
        artist = r.get("artist_name")
        text = r.get("relation_text_en") or ""
        if not artist or not text:
            kept.append(r)
            continue

        artist_group = member_group_map.get(_artist_key(canonical_artist(artist)))
        if not artist_group:
            kept.append(r)  # allowlist 밖 아티스트는 이미 앞 단계에서 걸러짐
            continue

        mentioned_groups = {
            member_group_map[_artist_key(name)]
            for name in extract_mentioned_names(text)
            if _artist_key(name) in member_group_map
        }
        if mentioned_groups and artist_group not in mentioned_groups:
            dropped.append((artist, text, sorted(mentioned_groups)))
        else:
            kept.append(r)
    return kept, dropped


def split_bilingual_name(name):
    """'Sushi Tatsu (스시타츠)' 같은 표기 -> (영문명, 한글명)."""
    if not name:
        return None, None
    m = re.match(r"^(.*?)\(([^)]+)\)\s*$", name.strip())
    if m:
        outer, inner = m.group(1).strip(), m.group(2).strip()
        if HANGUL_RE.search(inner) and not HANGUL_RE.search(outer):
            return (outer or None), inner
        if HANGUL_RE.search(outer) and not HANGUL_RE.search(inner):
            return inner, outer
    if HANGUL_RE.search(name):
        return None, name.strip()
    return name.strip(), None


def new_row(**kwargs):
    row = OrderedDict((f, None) for f in TEMPLATE_FIELDS)
    row.update(kwargs)
    return row


# ============================================================
# 4. 소스별 로더
# ============================================================

def load_tour_spots():
    """Kpop_Tour_Spots_Combined.xlsx - 가장 깨끗한 구조화 소스."""
    path = os.path.join(RAW_DIR, "Kpop_Tour_Spots_Combined.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["All K-pop Tour Spots"]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        artist = r[0]
        place = r[1] if len(r) > 1 else None
        cat = r[2] if len(r) > 2 else None
        loc = r[3] if len(r) > 3 else None
        notes = r[4] if len(r) > 4 else None
        src = r[5] if len(r) > 5 else None
        if not place:
            continue

        place_en, place_ko = split_bilingual_name(str(place).strip())
        city_id, city_ko, city_en, is_domestic, _ = parse_location(loc)
        mapped_cat = categorize(cat)

        review_reasons = []
        if city_id is None:
            review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님(해외 촬영/방문지 추정)")
        if mapped_cat is None:
            review_reasons.append(f"카테고리 미분류(원본: {cat!r})")

        rows.append({
            "artist_name": canonical_artist(str(artist).strip()),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "place_name_ko": place_ko, "place_name_en": place_en,
            "address": str(loc).strip() if loc else None,
            "relation_text_en": str(notes).strip() if notes else None,
            "place_category": mapped_cat,
            "source_url": str(src).strip() if src else None,
            "is_food": mapped_cat in FOOD_CATEGORIES if mapped_cat else None,
            "is_local_spot": None,
            "status": "needs_review" if review_reasons else "ready",
            "_review_reason": "; ".join(review_reasons),
            "_source": "Kpop_Tour_Spots_Combined.xlsx",
        })
    return rows


def load_thisismykorea():
    path = os.path.join(RAW_DIR, "stara_thisismykorea.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    for item in data:
        place = (item.get("place") or "").strip()
        if not place:
            continue

        review_reasons = []
        # 진짜 장소명이면 보통 12단어, 마침표 1개를 안 넘음. 그보다 길거나 문장이
        # 여러 개면 place/description 필드가 밀려서 설명문이 들어온 오염 행일
        # 가능성이 높음 (실제로 이 소스에서 2건 발견됨) -> 카카오 쿼리로 보내지 않음.
        if len(place.split()) > 12 or place.count(".") >= 2:
            review_reasons.append("place 필드가 설명문처럼 보임(필드 밀림 의심, 원본 데이터 확인 필요)")
            place_en, place_ko = None, None
        else:
            place_en, place_ko = split_bilingual_name(place)

        addr = item.get("address")
        # 실제 주소라면 번지수(숫자)가 거의 항상 포함됨. 숫자 없이 8단어 넘는 긴 문자열은
        # place/description/address 필드가 밀려 들어간 오염 행일 가능성이 높아 주소로 신뢰하지 않음.
        addr_text = str(addr or "")
        looks_like_address = bool(re.search(r"\d", addr_text)) or len(addr_text.split()) <= 8
        if not looks_like_address:
            city_id = city_ko = city_en = None
            review_reasons.append("주소 형식 아님(필드 밀림 의심, 원본 데이터 확인 필요)")
        else:
            city_id, city_ko, city_en, is_domestic, _ = parse_location(addr)
            if city_id is None:
                review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님")

        rows.append({
            "artist_name": canonical_artist((item.get("artist") or "").strip() or None),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "place_name_ko": place_ko, "place_name_en": place_en,
            "address": addr if looks_like_address else None,
            "relation_text_en": (item.get("description") or "").strip() or None,
            "source_url": item.get("url"),
            "is_food": None,
            "is_local_spot": None,
            "status": "needs_review" if review_reasons else "ready",
            "_review_reason": "; ".join(review_reasons),
            "_source": "stara_thisismykorea.json",
        })
    return rows


# 사람이 직접 검토하고 "이건 빼라"고 확정한 (아티스트, 장소명) 쌍.
# 예: "Chinatown" 검색은 지역 전체를 가리키는 이름이라 카카오에 단일 POI로
# 없고, 매번 그 이름이 들어간 딴 업체(주차장/게스트하우스 등)에 잘못 매칭됨 ->
# 자동 매칭 로직으로 못 고치는 케이스라 수동 제외.
MANUAL_EXCLUSIONS = {
    ("EXO", "chinatown"),
}


# "장소가 아닌 목차/질문/안내성 문구" 판별 - README가 요구한 "미분류/노이즈 필터"에 해당
NOISE_START_WORDS = (
    "why ", "how ", "is ", "are ", "where ", "what ", "understanding ",
    "best ", "sample ", "a full-day", "tax-free", "online ", "stadium / dome tour",
)


def clean_listicle_title(name):
    """'2. KWANGYA@SEOUL & SM Town COEX (Seoul Forest / Gangnam)' 같은 표기 정리.
    앞 번호 제거, ' — 부제'/' - 부제'는 관계 텍스트로 분리."""
    s = re.sub(r"^\d+\.\s*", "", name.strip())
    subtitle = None
    m = re.match(r"^(.*?)\s+[—–]\s+(.*)$", s)
    if m:
        s, subtitle = m.group(1).strip(), m.group(2).strip()
    return s, subtitle


def is_noise_title(name):
    s = name.strip().lower()
    if s.endswith("?"):
        return True
    return any(s.startswith(w) for w in NOISE_START_WORDS)


def load_stara_places():
    path = os.path.join(RAW_DIR, "stara_places.json")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    dropped = 0
    for item in data:
        raw_place = (item.get("place") or "").strip()
        if not raw_place:
            continue
        cleaned, subtitle = clean_listicle_title(raw_place)
        if is_noise_title(cleaned) or is_noise_title(raw_place):
            dropped += 1
            continue

        artists = item.get("artists") or []
        place_en, place_ko = split_bilingual_name(cleaned)

        # 이 소스는 주소/지역 필드가 아예 없음 -> city는 항상 비어있음.
        # (URL이 seoultourism.org 계열이라 실제로는 대부분 서울이겠지만, 문서 단위로
        # 확정된 정보가 아니라서 추측 대입하지 않고 사람 확인으로 넘김)
        review_reasons = ["city 정보 없음(소스에 주소 필드 없음)"]
        if not artists:
            review_reasons.append("아티스트 매칭 없음(노이즈 가능성, 사람 확인 필요)")

        base_desc = (item.get("description") or "").strip()
        relation = base_desc
        if subtitle:
            relation = (subtitle + (" - " + base_desc if base_desc else "")).strip()

        for artist in (artists or [None]):
            rows.append({
                "artist_name": canonical_artist(artist) if artist else None,
                "place_name_ko": place_ko, "place_name_en": place_en,
                "relation_text_en": relation or None,
                "image_url": item.get("image") or None,
                "source_url": item.get("url"),
                "is_food": None, "is_local_spot": None,
                "status": "needs_review",
                "_review_reason": "; ".join(review_reasons),
                "_source": "stara_places.json",
            })
    print(f"  stara_places.json: {len(data)}건 중 노이즈 {dropped}건 제외")
    return rows


def load_manual_places():
    """raw_data_정인지/*.json - 팀원이 직접 블로그 등에서 확인하고 정리한 장소.
    이미 정확한 주소/카테고리를 사람이 확인해서 넣은 소스라 다른 로더보다
    신뢰도가 높음 (category_hint는 우리 taxonomy로 이미 분류된 값이라
    categorize() 안 거치고 그대로 씀). 새 블로그를 추가하려면 이 폴더에
    같은 스키마의 json 파일을 추가하면 자동으로 인식됨."""
    import glob
    rows = []
    for path in glob.glob(os.path.join(BASE_DIR, "raw_data_정인지", "*.json")):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            place_ko = (item.get("place_name_ko") or "").strip()
            if not place_ko:
                continue
            city_id, city_ko, city_en, is_domestic, _ = parse_location(item.get("address"))

            review_reasons = []
            if item.get("needs_manual_check"):
                review_reasons.append("사람이 직접 표시한 확인 필요 항목(원본 데이터 내 주소 불일치 등)")
            if city_id is None:
                review_reasons.append("도시 매핑 실패" if is_domestic else "국내 지명 아님")

            cat = item.get("category_hint")
            rows.append({
                # 그룹 단위로 일관되게 (다른 소스들도 멤버 개별이 아니라 그룹명을
                # artist_name으로 씀; 멤버 정보는 description_ko/en 안에 이미 있음)
                "artist_name": canonical_artist(item.get("artist")),
                "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
                "place_name_ko": place_ko, "place_name_en": None,
                "address": item.get("address"),
                "relation_text_ko": item.get("description_ko"),
                "relation_text_en": item.get("description_en"),
                "place_category": cat,
                "source_url": item.get("url"),
                "is_food": cat in FOOD_CATEGORIES if cat else None,
                "is_local_spot": None,
                "status": "needs_review" if review_reasons else "ready",
                "_review_reason": "; ".join(review_reasons),
                "_source": os.path.basename(path),
            })
    return rows


def load_hometowns():
    """곽채원 xlsx + 정인지 CSV 합집합. (Idol, Group) 기준 중복 제거,
    누락 필드는 서로 보완."""
    combined = {}

    xlsx_path = os.path.join(RAW_DIR, "Kpop_Idol_Hometowns_filtered.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Idol Hometowns"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        idol, group, province, district = r[0], r[1], r[2], r[3]
        key = (str(idol).strip().lower(), str(group).strip().lower())
        combined[key] = {"idol": idol, "group": group, "province": province, "district": district}

    import glob
    csv_candidates = glob.glob(os.path.join(BASE_DIR, "아이돌_출신지_*.csv"))
    if not csv_candidates:
        raise FileNotFoundError("아이돌_출신지_*.csv 파일을 찾을 수 없음")
    csv_path = csv_candidates[0]
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # 빈 첫 줄
        header = next(reader)  # Idol,Group,Province,...
        for r in reader:
            if not r or not r[0]:
                continue
            idol, group, province, district = r[0], r[1], r[2], r[3]
            key = (str(idol).strip().lower(), str(group).strip().lower())
            existing = combined.get(key, {})
            combined[key] = {
                "idol": idol,
                "group": group,
                "province": province or existing.get("province"),
                "district": district or existing.get("district"),
            }

    rows = []
    for rec in combined.values():
        location = rec["district"] or rec["province"]
        city_id, city_ko, city_en, is_domestic, _ = parse_location(location)
        if city_id is None and rec["province"]:
            city_id, city_ko, city_en, is_domestic, _ = parse_location(rec["province"])

        rows.append({
            "artist_name": canonical_artist(str(rec["idol"]).strip()),
            "city_id": city_id, "city_name_ko": city_ko, "city_name_en": city_en,
            "relation_text_ko": f"{rec['idol']}의 고향",
            "relation_text_en": f"{rec['idol']}'s hometown",
            "quest_type": "hometown",
            "is_food": False,
            "is_local_spot": True,
            "status": "needs_review" if city_id is None else "ready",
            "_review_reason": "" if city_id is not None else "도시 매핑 실패",
            "_source": "hometowns(xlsx+csv 병합)",
            "_group": rec["group"],
        })
    return rows


# ============================================================
# 5. 병합(중복 제거) - place 계열만 대상. hometown은 애초에 (Idol,Group) 단위라 제외.
# ============================================================

def merge_places(rows):
    """같은 장소로 판단되는 행들을 묶되, relation_text/source_url처럼 "이 아티스트와
    이 장소의 관계"를 설명하는 필드는 아티스트별로 따로 유지함.
    예전엔 그룹의 대표 행(group[0]) 하나의 relation_text를 그룹 내 모든 아티스트에게
    그대로 복사해서 붙였는데, 실제로 같은 장소를 여러 아티스트가 각자 다른 이유로
    방문한 정당한 케이스(예: 파라다이스시티 - ENHYPEN 소속사 행사 vs BTS 뷔 앰버서더)에서
    서로 다른 relation_text가 뒤섞이는 문제가 있었음. place 자체의 공통 정보
    (이름/카테고리/이미지)만 그룹 전체에서 공유하고, 관계 설명은 아티스트 단위로 분리."""
    merged = []
    used = [False] * len(rows)
    for i, row in enumerate(rows):
        if used[i]:
            continue
        group = [row]
        used[i] = True
        name_i = row.get("place_name_en") or row.get("place_name_ko")
        for j in range(i + 1, len(rows)):
            if used[j]:
                continue
            other = rows[j]
            name_j = other.get("place_name_en") or other.get("place_name_ko")
            if row.get("city_id") == other.get("city_id") and is_same_place(name_i, name_j):
                group.append(other)
                used[j] = True

        shared = {}
        for field in ("place_name_ko", "place_name_en", "place_category", "image_url", "address"):
            for g in group:
                if g.get(field):
                    shared[field] = g[field]
                    break

        by_artist = OrderedDict()
        for g in group:
            by_artist.setdefault(g.get("artist_name"), []).append(g)

        for artist, artist_rows in by_artist.items():
            base = artist_rows[0]
            sources = sorted({g["source_url"] for g in artist_rows if g.get("source_url")})
            relations_en = sorted({g["relation_text_en"] for g in artist_rows if g.get("relation_text_en")})
            relations_ko = sorted({g["relation_text_ko"] for g in artist_rows if g.get("relation_text_ko")})
            reason_fragments = set()
            for g in artist_rows:
                for frag in (g.get("_review_reason") or "").split("; "):
                    if frag:
                        reason_fragments.add(frag)
            # "카테고리 미분류" 사유는 place 단위 속성이라, 그룹 내 다른 아티스트의
            # 행에서 카테고리가 이미 채워졌다면(shared) 이 아티스트에게도 더 이상
            # 유효한 사유가 아님 - 제거.
            if shared.get("place_category"):
                reason_fragments = {f for f in reason_fragments if not f.startswith("카테고리 미분류")}
            reasons = sorted(reason_fragments)

            merged_row = dict(base)
            merged_row.update(shared)
            merged_row["artist_name"] = artist
            merged_row["source_url"] = "; ".join(sources) if sources else base.get("source_url")
            merged_row["relation_text_en"] = " / ".join(relations_en) if relations_en else base.get("relation_text_en")
            merged_row["relation_text_ko"] = " / ".join(relations_ko) if relations_ko else base.get("relation_text_ko")
            merged_row["_review_reason"] = "; ".join(reasons)
            merged_row["status"] = "needs_review" if reasons else "ready"
            merged.append(merged_row)
    return merged


# ============================================================
# 6. ID 부여 + template 행 변환
# ============================================================

def finalize(rows, is_hometown=False):
    final_rows = []
    for r in rows:
        artist_name = r.get("artist_name")
        city_id = r.get("city_id")
        place_name = r.get("place_name_en") or r.get("place_name_ko")

        artist_id = f"art_{slugify(artist_name)}" if artist_name else None
        place_id = None if is_hometown else (
            f"pl_{r.get('city_id') or 'unk'}_{slugify(place_name)}" if place_name else None
        )

        row = new_row(
            artist_id=artist_id,
            artist_name=artist_name,
            city_id=city_id,
            city_name_ko=r.get("city_name_ko"),
            city_name_en=r.get("city_name_en"),
            place_id=place_id,
            place_name_ko=r.get("place_name_ko"),
            place_name_en=r.get("place_name_en"),
            relation_text_ko=r.get("relation_text_ko"),
            relation_text_en=r.get("relation_text_en"),
            place_category=r.get("place_category"),
            quest_type=r.get("quest_type"),
            image_url=r.get("image_url"),
            source_url=r.get("source_url"),
            is_food=r.get("is_food"),
            is_local_spot=r.get("is_local_spot"),
            status=r.get("status"),
        )
        row["_review_reason"] = r.get("_review_reason", "")
        row["_source"] = r.get("_source", "")
        row["_address"] = r.get("address", "")
        final_rows.append(row)
    return final_rows


# ============================================================
# 7. 지오코딩 + 한글 상호명 채우기 (Kakao 키워드 장소검색)
# ============================================================
#
# 주소 검색이 아니라 "키워드(장소명) 검색"을 쓰는 이유: 이 데이터의 place_name은
# 거의 다 영문(관광 사이트에서 긁어온 것)이고, 원본 street address도 template에는
# 남겨두지 않았음(city 단위로만 정규화함). 반면 키워드 검색은 장소명 텍스트로
# 카카오 로컬DB를 찾아서, 매칭되면 "실제 등록된 한글 상호명"과 좌표를 동시에
# 돌려줌 -> place_name_ko를 LLM으로 추측 번역하는 것보다 훨씬 신뢰도 높음.
# 매칭 안 되면(오탐 방지 위해 도시 불일치 매칭도 버림) 그냥 비워두고 needs_review.

CACHE_PATH = os.path.join(OUT_DIR, ".geocode_cache.json")


def load_env():
    """STARA/.env를 읽어서 os.environ에 반영 (python-dotenv 의존성 없이 최소 구현)."""
    env_path = os.path.join(BASE_DIR, ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key, val = key.strip(), val.strip()
            if key and key not in os.environ:
                os.environ[key] = val


def load_geocode_cache():
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_geocode_cache(cache):
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def kakao_keyword_search(query, api_key):
    resp = requests.get(
        "https://dapi.kakao.com/v2/local/search/keyword.json",
        params={"query": query, "size": 5},
        headers={"Authorization": f"KakaoAK {api_key}"},
        timeout=10,
    )
    resp.raise_for_status()
    return resp.json().get("documents", [])


# 관광/방문 목적지가 될 수 없는 카테고리 - 이런 매칭은 이름만 비슷하고 실제로는
# 전혀 다른 시설인 경우가 많음 (예: "Chinatown" 검색이 "차이나타운 공영주차장"에
# 매칭된 사례, "Citizen Park" 검색이 근처 아파트 단지에 매칭된 사례 실제 확인됨)
NOISE_CATEGORIES = ("주차장", "부동산", "아파트", "주거시설")


def pick_best_match(docs, city_ko):
    """도시명이 결과 주소에 포함된 것만 신뢰 (엉뚱한 지역의 동명 장소 오매칭 방지).
    주의: 이것만으로는 부족함 - "Busan Citizen Park" 검색이 "시민공원아파트"(아파트
    단지)에 매칭되는 것처럼, 도시만 맞고 실제로는 다른 장소인 오탐이 확인됨.
    그래서 이 결과를 쓰는 쪽(geocode_and_fill_names)에서 영문 일반명사성 쿼리는
    무조건 needs_review로 내려서 사람이 좌표를 눈으로 확인하게 함."""
    docs = [d for d in docs if not any(nc in (d.get("category_name") or "") for nc in NOISE_CATEGORIES)]
    if not docs:
        return None
    if not city_ko:
        return docs[0]
    city_key = city_ko.replace("특별시", "").replace("광역시", "").replace("특별자치시", "").replace("특별자치도", "").replace("도", "")
    for doc in docs:
        addr = (doc.get("road_address_name") or "") + (doc.get("address_name") or "")
        if city_key and city_key in addr:
            return doc
    return None  # 결과는 있지만 다 다른 도시 -> 신뢰 안 함


# Kakao category_name(예: "음식점 > 한식 > 육류,고기")을 우리 place_category taxonomy로 매핑
KAKAO_CATEGORY_RULES = [
    (("음식점",), "음식점"),
    (("카페", "제과,베이커리", "디저트"), "카페/디저트"),
    (("쇼핑", "판매", "마트", "백화점", "패션"), "쇼핑"),
    (("박물관", "전시관", "미술관"), "박물관/전시"),
    (("테마파크",), "테마파크"),
    (("공원",), "공원"),
    (("해수욕장", "해변"), "해변"),
    (("고궁", "유적", "사적"), "유적지"),
    (("공연장", "콘서트홀", "경기장", "운동장"), "공연장"),
    (("숙박", "호텔", "모텔", "펜션", "리조트"), "숙박"),
    (("관광,명소", "랜드마크", "타워"), "랜드마크"),
    (("사진", "포토"), "포토스팟"),
    (("찜질방", "사우나"), "체험/휴양"),
    (("스포츠", "레저", "액티비티", "낚시"), "체험/액티비티"),
]


def categorize_kakao(category_name):
    if not category_name:
        return None
    for keywords, mapped in KAKAO_CATEGORY_RULES:
        if any(kw in category_name for kw in keywords):
            return mapped
    return None


def geocode_and_fill_names(rows, api_key):
    cache = load_geocode_cache()
    n_queried = n_hit = n_cached = n_skipped = 0

    targets = [
        r for r in rows
        if (r.get("place_name_ko") or r.get("place_name_en"))
        and (r.get("latitude") is None or r.get("place_name_ko") is None)
    ]
    print(f"   지오코딩 대상: {len(targets)}행 (전체 {len(rows)}행 중 장소명 있는 것)")

    for i, row in enumerate(targets, 1):
        had_korean_name = bool(row.get("place_name_ko"))
        name = row.get("place_name_ko") or row.get("place_name_en")
        city_en = row.get("city_name_en") or ""
        query = f"{city_en} {name}".strip()

        if query in cache:
            docs = cache[query]
            n_cached += 1
        else:
            try:
                docs = kakao_keyword_search(query, api_key)
            except requests.RequestException as e:
                # Windows 콘솔(cp949)이 원본 텍스트의 특수문자를 못 그리는 경우가 있어서
                # 에러 로그 출력 자체가 죽지 않게 ascii로 안전하게 치환해서 출력.
                safe_query = query.encode("ascii", "replace").decode("ascii")
                print(f"   [실패] {safe_query!r} -> {e}")
                docs = []
            cache[query] = docs
            n_queried += 1
            time.sleep(0.15)  # 초당 요청 제한 여유

        if i % 50 == 0:
            print(f"   ...{i}/{len(targets)} 처리, 신규 조회 {n_queried}건")

        best = pick_best_match(docs, row.get("city_name_ko"))
        if not best:
            frag = "카카오 키워드 검색 매칭 없음(이름/좌표 확인 필요)"
            existing = row.get("_review_reason") or ""
            if frag not in existing:
                row["_review_reason"] = (existing + "; " + frag).strip("; ")
            row["status"] = "needs_review"
            continue

        n_hit += 1
        if not row.get("place_name_ko"):
            row["place_name_ko"] = best.get("place_name")
        row["latitude"] = float(best["y"])
        row["longitude"] = float(best["x"])

        # city가 원래 없었으면(주소 필드 자체가 없던 stara_places.json 등),
        # 카카오가 돌려준 실제 주소를 우리 parse_location으로 파싱해서 채움.
        if not row.get("city_id"):
            addr = best.get("road_address_name") or best.get("address_name")
            city_id, city_ko, city_en, _, _ = parse_location(addr)
            if city_id:
                row["city_id"], row["city_name_ko"], row["city_name_en"] = city_id, city_ko, city_en
                fragments = [f for f in (row.get("_review_reason") or "").split("; ")
                             if f and not f.startswith("도시 매핑 실패")
                             and not f.startswith("city 정보 없음")
                             and not f.startswith("국내 지명 아님")]
                row["_review_reason"] = "; ".join(fragments)

        # 카테고리가 이걸로 새로 채워졌으면, 로딩 단계에서 붙었던
        # "카테고리 미분류(...)" 사유는 더 이상 유효하지 않으니 제거.
        if not row.get("place_category"):
            mapped_cat = categorize_kakao(best.get("category_name"))
            if mapped_cat:
                row["place_category"] = mapped_cat
                row["is_food"] = mapped_cat in FOOD_CATEGORIES
                fragments = [f for f in (row.get("_review_reason") or "").split("; ")
                             if f and not f.startswith("카테고리 미분류")]
                row["_review_reason"] = "; ".join(fragments)

        if not had_korean_name:
            # 원래 영문(설명형) 이름으로 검색한 매칭은 "도시가 같다"만 확인됐을 뿐 실제
            # 동일 장소인지는 검증이 약함(예: "Citizen Park" -> 근처 아파트단지가
            # 걸린 사례 실제 확인됨). 좌표는 채워두되 사람이 한 번 봐야 함.
            frag = "영문명으로 검색한 매칭(실제 동일 장소인지 확인 필요)"
            existing = row.get("_review_reason") or ""
            if frag not in existing:
                row["_review_reason"] = (existing + "; " + frag).strip("; ")

        # 남은 사유가 없으면 ready로 승격, 있으면 needs_review 유지.
        row["status"] = "needs_review" if row.get("_review_reason") else "ready"

    save_geocode_cache(cache)
    print(f"   완료: 신규조회 {n_queried} / 캐시재사용 {n_cached} / 매칭성공 {n_hit} / 대상 {len(targets)}")


# ============================================================
# 8. 저장
# ============================================================

def save_outputs(rows):
    os.makedirs(OUT_DIR, exist_ok=True)

    json_rows = [{k: v for k, v in r.items() if k in TEMPLATE_FIELDS} for r in rows]
    with open(os.path.join(OUT_DIR, "dataset.json"), "w", encoding="utf-8") as f:
        json.dump(json_rows, f, ensure_ascii=False, indent=2)

    with open(os.path.join(OUT_DIR, "dataset.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_FIELDS)
        writer.writeheader()
        for r in json_rows:
            writer.writerow(r)

    review_fields = TEMPLATE_FIELDS + ["_source", "_review_reason"]
    review_rows = [r for r in rows if r.get("status") == "needs_review"]
    with open(os.path.join(OUT_DIR, "needs_review.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=review_fields)
        writer.writeheader()
        for r in review_rows:
            writer.writerow({k: r.get(k) for k in review_fields})

    return len(json_rows), len(review_rows)


# ============================================================
# 9. 새 template(Data_Preprocessing_Template.ts, Place interface) 변환
# ============================================================
#
# 기존 template과 구조가 완전히 다름: city 필드 전부 삭제, artistIds가 배열로
# 바뀜(장소 하나에 여러 아티스트) - 지금까지는 아티스트별로 행을 나눴는데
# 반대로 장소 단위로 다시 묶어야 함. category도 15종 -> 5종으로 축소됨.
# openTime/closeTime/dwellMinutes는 인터페이스상 필수(옵셔널 아님)인데 우리가
# 수집한 적 없는 데이터라 빈 값으로 둠 - 실제 운영시간 조사가 별도로 필요함.

CATEGORY_TO_TS = {
    "음식점": "food", "카페/디저트": "food",
    "쇼핑": "shopping",
    "촬영지": "photo", "공원": "photo", "해변": "photo",
    "랜드마크": "photo", "포토스팟": "photo",
    "박물관/전시": "culture", "유적지": "culture",
    "기획사": "culture", "공연장": "culture",
    "테마파크": "experience", "숙박": "experience",
    "체험/액티비티": "experience", "체험/휴양": "experience",
}


def transform_to_new_template(rows):
    """장소 단위(place_id)로 다시 묶어서 새 Place 인터페이스 형태로 변환.
    같은 장소를 여러 아티스트가 공유하면 artistIds에 다 모으고,
    relationTextKo/En도 아티스트별 설명을 합쳐서 하나로 만듦."""
    by_place = OrderedDict()
    for r in rows:
        pid = r.get("place_id")
        if not pid:
            continue
        by_place.setdefault(pid, []).append(r)

    result = []
    for pid, group in by_place.items():
        base = group[0]
        name_ko = base.get("place_name_ko") or ""
        name_en = base.get("place_name_en") or ""
        place_slug = slugify_hyphen(name_en or name_ko)

        artist_slugs = []
        for g in group:
            slug = slugify_hyphen(g.get("artist_name"))
            if slug and slug not in artist_slugs:
                artist_slugs.append(slug)
        primary_artist = artist_slugs[0] if artist_slugs else "unknown"

        cat_ko = base.get("place_category")
        cat_en = CATEGORY_TO_TS.get(cat_ko, "experience")

        relations_ko = [g["relation_text_ko"] for g in group if g.get("relation_text_ko")]
        relations_en = [g["relation_text_en"] for g in group if g.get("relation_text_en")]

        lat = base.get("latitude")
        lon = base.get("longitude")
        image_url = next((g.get("image_url") for g in group if g.get("image_url")), None)
        address = next((g.get("_address") for g in group if g.get("_address")), None)
        is_food = any(g.get("is_food") for g in group)
        statuses = {g.get("status") for g in group}

        result.append({
            "id": f"{primary_artist}-{cat_en}-{place_slug}",
            "nameKo": name_ko,
            "nameEn": name_en,
            "latitude": lat,
            "longitude": lon,
            "category": cat_en,
            "artistIds": artist_slugs,
            "relationTextKo": " / ".join(dict.fromkeys(relations_ko)),
            "relationTextEn": " / ".join(dict.fromkeys(relations_en)),
            "openTime": "",
            "closeTime": "",
            "dwellMinutes": None,
            "imageUrl": image_url,
            "isFood": is_food,
            "isLocalSpot": False,
            "isMainRoute": False,
            "address": address,
            "_status": "needs_review" if "needs_review" in statuses else "ready",
        })
    return result


def save_rows_as(subset, label):
    """이미 걸러진 행 목록을 label_dataset.csv/json + label_needs_review.csv로 저장."""
    json_rows = [{k: v for k, v in r.items() if k in TEMPLATE_FIELDS} for r in subset]

    with open(os.path.join(OUT_DIR, f"{label}_dataset.json"), "w", encoding="utf-8") as f:
        json.dump(json_rows, f, ensure_ascii=False, indent=2)

    with open(os.path.join(OUT_DIR, f"{label}_dataset.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_FIELDS)
        writer.writeheader()
        for r in json_rows:
            writer.writerow(r)

    review_fields = TEMPLATE_FIELDS + ["_source", "_review_reason"]
    review_rows = [r for r in subset if r.get("status") == "needs_review"]
    with open(os.path.join(OUT_DIR, f"{label}_needs_review.csv"), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=review_fields)
        writer.writeheader()
        for r in review_rows:
            writer.writerow({k: r.get(k) for k in review_fields})

    return len(subset), len(review_rows)


def main():
    print("1) Kpop_Tour_Spots_Combined.xlsx 로딩...")
    tour_rows = load_tour_spots()
    print(f"   {len(tour_rows)}행")

    print("2) stara_thisismykorea.json 로딩...")
    tmk_rows = load_thisismykorea()
    print(f"   {len(tmk_rows)}행")

    print("3) stara_places.json 로딩 (노이즈 필터 적용)...")
    sp_rows = load_stara_places()
    print(f"   {len(sp_rows)}행")

    print("3b) raw_data_정인지/*.json 로딩 (직접 확인한 블로그 등)...")
    manual_rows = load_manual_places()
    print(f"   {len(manual_rows)}행")

    print("4) 아이돌 출신지 (xlsx + CSV) 병합...")
    home_rows = load_hometowns()
    print(f"   {len(home_rows)}행")

    print("5) 장소 데이터(1~3) 중복 병합...")
    place_rows = merge_places(tour_rows + tmk_rows + sp_rows + manual_rows)
    print(f"   병합 전 {len(tour_rows) + len(tmk_rows) + len(sp_rows) + len(manual_rows)}행 -> 병합 후 {len(place_rows)}행")

    place_rows = filter_manual_exclusions(place_rows)

    print("5b) 아티스트 allowlist(artist_allowlist.json) 필터링...")
    allowed_keys = load_artist_allowlist()
    place_rows, dropped_artists = filter_by_allowlist(place_rows, allowed_keys)
    print(f"   리스트 밖 아티스트 {len(dropped_artists)}종 제외: {sorted(dropped_artists)}")
    print(f"   {len(place_rows)}행 남음")

    print("5c) artist_name vs relation_text 인물 불일치 필터링...")
    member_group_map = load_member_group_map()
    place_rows, dropped_mismatches = filter_by_relation_consistency(place_rows, member_group_map)
    print(f"   불일치로 {len(dropped_mismatches)}행 제외")
    for artist, text, groups in dropped_mismatches[:20]:
        safe_text = text.encode("ascii", "replace").decode("ascii")
        print(f"     - artist={artist!r} vs 텍스트 속 그룹={groups} | {safe_text[:70]!r}")
    print(f"   {len(place_rows)}행 남음")

    print("6) template 스키마로 변환 + ID 부여...")
    final_places = finalize(place_rows, is_hometown=False)
    final_homes = finalize(home_rows, is_hometown=True)
    all_rows = final_places + final_homes

    if "--geocode" in sys.argv:
        print("7) 지오코딩 + 한글 상호명 채우기 (Kakao 키워드 검색)...")
        load_env()
        api_key = os.environ.get("KAKAO_REST_API_KEY", "").strip()
        if not api_key:
            print("   [건너뜀] KAKAO_REST_API_KEY가 .env에 없음")
        else:
            geocode_and_fill_names(all_rows, api_key)
    else:
        print("7) 지오코딩 건너뜀 (실행하려면: python build_dataset.py --geocode)")

    print("8) 저장...")
    total, review = save_outputs(all_rows)
    print(f"\n완료: 총 {total}행 -> preprocessed/dataset.csv, dataset.json")
    print(f"검수 필요: {review}행 -> preprocessed/needs_review.csv")
    ready = total - review
    print(f"(ready: {ready} / needs_review: {review})")

    print("\n9) MVP 범위(인천) 추출...")
    # 인천 MVP는 출신지(고향) 데이터 제외 - 실제 방문 가능한 장소 퀘스트만 다룸.
    # is_local_spot도 이 범위에선 전부 False로 통일 (인천 MVP 단계에서는
    # "로컬 명소" 구분을 아직 안 쓰기로 함).
    incheon_rows = [r for r in all_rows if r.get("city_id") == "incheon" and r.get("quest_type") != "hometown"]
    for r in incheon_rows:
        r["is_local_spot"] = False
    incheon_total, incheon_review = save_rows_as(incheon_rows, "incheon")
    print(f"인천: 총 {incheon_total}행 (ready {incheon_total - incheon_review} / needs_review {incheon_review})")
    print("-> preprocessed/incheon_dataset.csv, incheon_dataset.json, incheon_needs_review.csv")

    print("\n10) 새 template(Place 인터페이스)로 변환 - 인천만...")
    new_places = transform_to_new_template(incheon_rows)
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "incheon_places.json"), "w", encoding="utf-8") as f:
        json.dump([{k: v for k, v in p.items() if k != "_status"} for p in new_places],
                   f, ensure_ascii=False, indent=2)
    needs_review_new = sum(1 for p in new_places if p["_status"] == "needs_review")
    print(f"장소(place) 단위로 재구성: {len(new_places)}개 (아티스트별 행 -> 장소별 행, artistIds 배열)")
    print(f"needs_review: {needs_review_new}개")
    print("-> preprocessed/incheon_places.json (Place[] 형태, 기존 incheon_dataset.json은 그대로 둠)")


if __name__ == "__main__":
    main()
